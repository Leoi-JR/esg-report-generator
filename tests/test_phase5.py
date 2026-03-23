"""
tests/test_phase5.py
====================
阶段五：输出对齐表 Excel 专项测试。

测试策略：使用 tempfile 临时目录写入 Excel，openpyxl 读回验证。
Mock alignment_records + indicator_details，不依赖真实数据 / API。

运行方式：
    conda run -n esg python3 tests/test_phase5.py
"""

import os
import sys
import io
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from align_evidence import (
    write_alignment_excel,
    print_phase5_summary,
    ALIGNMENT_COLUMNS,
    CONSISTENCY_SORT_ORDER,
)


# ==============================================================================
# 辅助：构造 mock 数据
# ==============================================================================

def _mock_indicator_details():
    """返回少量编码映射（GA1, EB1, SC2 等）。"""
    return {
        "GA1": {
            "topic": "可持续发展治理机制",
            "indicator": "可持续发展目标与愿景",
            "requirement": "请提供相关文件",
        },
        "EB1": {
            "topic": "环境管理体系",
            "indicator": "环保目标与管理",
            "requirement": "环保管理制度",
        },
        "SC2": {
            "topic": "劳工权益",
            "indicator": "员工权益保障",
            "requirement": "员工手册",
        },
    }


def _mock_alignment_records(
    n: int = 5,
    statuses: list = None,
    folder_codes: list = None,
):
    """
    生成含全部字段的 mock 对齐记录。

    参数：
        n            - 记录数量
        statuses     - 每条记录的 consistency 状态（长度 == n）
        folder_codes - 每条记录的 folder_code（长度 == n）
    """
    if statuses is None:
        statuses = ["✅"] * n
    if folder_codes is None:
        folder_codes = ["GA1"] * n

    records = []
    for i in range(n):
        records.append({
            "chunk_id":        f"file{i}.pdf#s0#c0",
            "file_path":       f"/mock/target/G-公司治理/GA1/file{i}.pdf",
            "file_name":       f"file{i}.pdf",
            "folder_code":     folder_codes[i],
            "page_or_sheet":   "1",
            "chunk_index":     i,
            "text":            f"这是第 {i} 条测试文本。" * 10,
            "parent_text":     f"完整文本 {i}。" * 20,
            "char_count":      100,
            "embedding":       [0.1] * 8,
            "semantic_topk":   [("GA1", 0.91), ("GA2", 0.87), ("EB1", 0.72)],
            "consistency":     statuses[i],
            "consistency_desc": "测试描述",
            "suggested_code":  folder_codes[i] if statuses[i] in ("✅", "➕") else "EB1",
        })
    return records


# ==============================================================================
# TestExcelStructure（5 个用例）
# ==============================================================================

class TestExcelStructure(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        """创建临时目录，写入一次 Excel 供本组用例共享。"""
        cls.tmpdir = tempfile.mkdtemp()
        cls.records = _mock_alignment_records(n=10, statuses=["✅"] * 10)
        cls.details = _mock_indicator_details()
        cls.output_path = write_alignment_excel(
            cls.records, cls.details, cls.tmpdir, "/mock/target", "测试公司",
        )
        from openpyxl import load_workbook
        cls.wb = load_workbook(cls.output_path)
        cls.ws = cls.wb.active

    def test_column_count(self):
        """列数 == 13。"""
        self.assertEqual(self.ws.max_column, 13)

    def test_column_names_match(self):
        """列名与 ALIGNMENT_COLUMNS 完全一致。"""
        headers = [cell.value for cell in self.ws[1]]
        self.assertEqual(headers, ALIGNMENT_COLUMNS)

    def test_row_count_matches_input(self):
        """数据行数 == 输入记录数。"""
        # max_row 含表头行，数据行 = max_row - 1
        data_rows = self.ws.max_row - 1
        self.assertEqual(data_rows, len(self.records))

    def test_empty_input_creates_header_only(self):
        """空列表输入 → 只有表头行。"""
        tmpdir = tempfile.mkdtemp()
        path = write_alignment_excel([], _mock_indicator_details(), tmpdir, "/mock", "测试")
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        # 只有 1 行（表头）
        self.assertEqual(ws.max_row, 1)
        # 列名仍完整
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ALIGNMENT_COLUMNS)

    def test_returns_valid_path(self):
        """返回路径存在且以 .xlsx 结尾。"""
        self.assertTrue(os.path.exists(self.output_path))
        self.assertTrue(self.output_path.endswith(".xlsx"))


# ==============================================================================
# TestSorting（2 个用例）
# ==============================================================================

class TestSorting(unittest.TestCase):

    def test_sort_order(self):
        """⚠️ 行在 🔍 之前，🔍 在 ➕ 之前，➕ 在 ✅ 之前，✅ 在 ❓ 之前。"""
        statuses = ["✅", "❓", "⚠️", "🔍", "➕"]
        records = _mock_alignment_records(n=5, statuses=statuses)
        tmpdir = tempfile.mkdtemp()
        path = write_alignment_excel(
            records, _mock_indicator_details(), tmpdir, "/mock/target", "测试",
        )
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        consistency_col = ALIGNMENT_COLUMNS.index("consistency") + 1
        actual_statuses = [
            ws.cell(row=r, column=consistency_col).value
            for r in range(2, ws.max_row + 1)
        ]
        expected_order = ["⚠️", "🔍", "➕", "✅", "❓"]
        self.assertEqual(actual_statuses, expected_order)

    def test_stable_sort_within_group(self):
        """同状态的行保持原始相对顺序。"""
        statuses = ["✅", "✅", "⚠️", "⚠️"]
        # folder_codes 用于区分行
        folder_codes = ["GA1", "EB1", "SC2", "GA1"]
        records = _mock_alignment_records(n=4, statuses=statuses, folder_codes=folder_codes)
        tmpdir = tempfile.mkdtemp()
        path = write_alignment_excel(
            records, _mock_indicator_details(), tmpdir, "/mock/target", "测试",
        )
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        # ⚠️ 排前面，两条 ⚠️ 应保持 SC2, GA1 的原始顺序
        folder_col = ALIGNMENT_COLUMNS.index("folder_code") + 1
        row2_folder = ws.cell(row=2, column=folder_col).value  # 第一条 ⚠️
        row3_folder = ws.cell(row=3, column=folder_col).value  # 第二条 ⚠️
        self.assertEqual(row2_folder, "SC2")
        self.assertEqual(row3_folder, "GA1")
        # ✅ 排后面，两条 ✅ 应保持 GA1, EB1 的原始顺序
        row4_folder = ws.cell(row=4, column=folder_col).value
        row5_folder = ws.cell(row=5, column=folder_col).value
        self.assertEqual(row4_folder, "GA1")
        self.assertEqual(row5_folder, "EB1")


# ==============================================================================
# TestFieldValues（9 个用例）
# ==============================================================================

class TestFieldValues(unittest.TestCase):

    def _write_and_read(self, records, details=None):
        """辅助：写入 Excel 并用 openpyxl 读回。"""
        if details is None:
            details = _mock_indicator_details()
        tmpdir = tempfile.mkdtemp()
        path = write_alignment_excel(records, details, tmpdir, "/mock/target", "测试")
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        return ws

    def test_semantic_top5_formatted(self):
        """[(GA1,0.91),(GA2,0.87)] → "GA1:0.91, GA2:0.87"。"""
        records = _mock_alignment_records(n=1)
        records[0]["semantic_topk"] = [("GA1", 0.91), ("GA2", 0.87)]
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("semantic_top5") + 1
        val = ws.cell(row=2, column=col).value
        self.assertEqual(val, "GA1:0.91, GA2:0.87")

    def test_semantic_top5_empty(self):
        """[] → 空字符串。"""
        records = _mock_alignment_records(n=1)
        records[0]["semantic_topk"] = []
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("semantic_top5") + 1
        val = ws.cell(row=2, column=col).value
        # openpyxl 可能读回 None 或 ""
        self.assertIn(val, [None, ""])

    def test_text_preview_truncated_200(self):
        """300 字文本 → preview 恰好 200 字。"""
        records = _mock_alignment_records(n=1)
        records[0]["text"] = "甲" * 300
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("text_preview") + 1
        val = ws.cell(row=2, column=col).value
        self.assertEqual(len(val), 200)

    def test_text_preview_newlines_replaced(self):
        """\n 和 \r 替换为空格。"""
        records = _mock_alignment_records(n=1)
        records[0]["text"] = "第一行\n第二行\r第三行"
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("text_preview") + 1
        val = ws.cell(row=2, column=col).value
        self.assertNotIn("\n", val)
        self.assertNotIn("\r", val)
        self.assertIn("第一行 第二行 第三行", val)

    def test_folder_topic_lookup(self):
        """folder_code=GA1 → 正确查表得到 topic。"""
        records = _mock_alignment_records(n=1, folder_codes=["GA1"])
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("folder_topic") + 1
        val = ws.cell(row=2, column=col).value
        self.assertEqual(val, "可持续发展治理机制")

    def test_folder_topic_none_code(self):
        """folder_code=None → 空字符串。"""
        records = _mock_alignment_records(n=1, folder_codes=[None])
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("folder_topic") + 1
        val = ws.cell(row=2, column=col).value
        self.assertIn(val, [None, ""])

    def test_folder_topic_unknown_code(self):
        """folder_code 不在 indicator_details → 空字符串。"""
        records = _mock_alignment_records(n=1, folder_codes=["ZZ99"])
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("folder_topic") + 1
        val = ws.cell(row=2, column=col).value
        self.assertIn(val, [None, ""])

    def test_human_code_prefilled(self):
        """suggested_code=EB1 → human_code=EB1。"""
        records = _mock_alignment_records(n=1, statuses=["⚠️"])
        records[0]["suggested_code"] = "EB1"
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("human_code") + 1
        val = ws.cell(row=2, column=col).value
        self.assertEqual(val, "EB1")

    def test_file_path_is_relative(self):
        """file_path 列存的是相对路径（非绝对路径）。"""
        records = _mock_alignment_records(n=1)
        records[0]["file_path"] = "/mock/target/G-公司治理/GA1/test.pdf"
        ws = self._write_and_read(records)
        col = ALIGNMENT_COLUMNS.index("file_path") + 1
        val = ws.cell(row=2, column=col).value
        # 相对路径不应以 / 开头
        self.assertFalse(val.startswith("/"), f"路径应为相对路径，实际为: {val}")


# ==============================================================================
# TestExcelFormatting（5 个用例）
# ==============================================================================

class TestExcelFormatting(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        """创建含多种状态的 Excel 供本组用例共享。"""
        cls.tmpdir = tempfile.mkdtemp()
        statuses = ["⚠️", "🔍", "✅", "✅", "➕"]
        cls.records = _mock_alignment_records(n=5, statuses=statuses)
        cls.details = _mock_indicator_details()
        cls.output_path = write_alignment_excel(
            cls.records, cls.details, cls.tmpdir, "/mock/target", "测试",
        )
        from openpyxl import load_workbook
        cls.wb = load_workbook(cls.output_path)
        cls.ws = cls.wb.active

    def test_freeze_panes(self):
        """ws.freeze_panes == "A2"。"""
        self.assertEqual(self.ws.freeze_panes, "A2")

    def test_warning_rows_red_fill(self):
        """⚠️ 行的单元格有浅红 fill。"""
        consistency_col = ALIGNMENT_COLUMNS.index("consistency") + 1
        for row_idx in range(2, self.ws.max_row + 1):
            cell_val = str(self.ws.cell(row=row_idx, column=consistency_col).value or "")
            if "⚠️" in cell_val:
                # 检查该行所有列的 fill
                for col_idx in range(1, self.ws.max_column + 1):
                    fill = self.ws.cell(row=row_idx, column=col_idx).fill
                    self.assertEqual(
                        fill.fgColor.rgb, "00FFC7CE",
                        f"⚠️ 行 row={row_idx} col={col_idx} 应有浅红背景"
                    )

    def test_search_rows_yellow_fill(self):
        """🔍 行的单元格有浅黄 fill。"""
        consistency_col = ALIGNMENT_COLUMNS.index("consistency") + 1
        for row_idx in range(2, self.ws.max_row + 1):
            cell_val = str(self.ws.cell(row=row_idx, column=consistency_col).value or "")
            if "🔍" in cell_val:
                for col_idx in range(1, self.ws.max_column + 1):
                    fill = self.ws.cell(row=row_idx, column=col_idx).fill
                    self.assertEqual(
                        fill.fgColor.rgb, "00FFEB9C",
                        f"🔍 行 row={row_idx} col={col_idx} 应有浅黄背景"
                    )

    def test_consistent_rows_no_fill(self):
        """✅ 行无特殊 fill。"""
        consistency_col = ALIGNMENT_COLUMNS.index("consistency") + 1
        for row_idx in range(2, self.ws.max_row + 1):
            cell_val = str(self.ws.cell(row=row_idx, column=consistency_col).value or "")
            if cell_val == "✅":
                for col_idx in range(1, self.ws.max_column + 1):
                    fill = self.ws.cell(row=row_idx, column=col_idx).fill
                    # 无特殊 fill → fgColor 应为默认值（00000000）或 None
                    fg_rgb = fill.fgColor.rgb if fill.fgColor else "00000000"
                    self.assertNotIn(
                        fg_rgb, ["00FFC7CE", "00FFEB9C"],
                        f"✅ 行不应有红/黄背景 row={row_idx} col={col_idx}"
                    )

    def test_column_width_capped(self):
        """所有列宽 ≤ 62（60 + 余量）。"""
        for col_letter, dim in self.ws.column_dimensions.items():
            self.assertLessEqual(
                dim.width, 62,
                f"列 {col_letter} 宽度 {dim.width} 超过上限 62"
            )


# ==============================================================================
# TestPrintSummary（1 个用例）
# ==============================================================================

class TestPrintSummary(unittest.TestCase):

    def test_summary_output(self):
        """捕获 stdout 包含"已生成"和"需人工审查"且计数正确。"""
        statuses = ["⚠️", "🔍", "➕", "✅", "✅", "❓"]
        records = _mock_alignment_records(n=6, statuses=statuses)
        # ⚠️ + 🔍 + ➕ = 3

        captured = io.StringIO()
        import contextlib
        with contextlib.redirect_stdout(captured):
            print_phase5_summary("/tmp/test.xlsx", records)

        output = captured.getvalue()
        self.assertIn("已生成", output)
        self.assertIn("需人工审查", output)
        self.assertIn("3", output, "⚠️ + 🔍 + ➕ = 3 行应出现在输出中")


# ==============================================================================
# 运行
# ==============================================================================

def run_all():
    """运行所有阶段五测试，打印摘要。"""
    suites = [
        unittest.TestLoader().loadTestsFromTestCase(TestExcelStructure),
        unittest.TestLoader().loadTestsFromTestCase(TestSorting),
        unittest.TestLoader().loadTestsFromTestCase(TestFieldValues),
        unittest.TestLoader().loadTestsFromTestCase(TestExcelFormatting),
        unittest.TestLoader().loadTestsFromTestCase(TestPrintSummary),
    ]
    combined = unittest.TestSuite(suites)
    runner   = unittest.TextTestRunner(verbosity=2)
    result   = runner.run(combined)
    return result


if __name__ == "__main__":
    run_all()
