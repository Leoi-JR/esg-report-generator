"""
esg_utils.py
============
ESG 工具链公共模块：存放所有脚本共用的常量、正则表达式和工具函数。

本模块由以下脚本共同引用：
  - generate_folder_structure.py
  - data_list_v2.py
  - simulate_client_sorting.py
  - align_evidence.py（待开发）

修改本模块中的函数后，所有引用脚本同步生效，无需逐一修改。
"""

import re
import warnings

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection

warnings.filterwarnings("ignore", category=UserWarning)


# ==============================================================================
# 公共常量
# ==============================================================================

# ESG 编码正则：A/G/E/D/S + 0~3 位字母 + 1~3 位数字，确保数字后不继续跟数字
CODE_REGEX = re.compile(
    r"(?<![A-Z0-9])([AGEDS][A-Z]{0,3}\d{1,3})(?!\d)", re.IGNORECASE
)

# 维度前缀 → 一级目录名 / 描述
DIMENSION_META = {
    "A": {"folder": "A-总体概况",   "desc": "总体概况：公司基本信息、发展历程、荣誉奖项等"},
    "G": {"folder": "G-公司治理",   "desc": "公司治理：可持续发展治理、规范治理与党建、商业行为"},
    "E": {"folder": "E-环境保护",   "desc": "环境保护：应对气候变化、污染物管控、资源利用、生态多样性"},
    "D": {"folder": "D-产业价值",   "desc": "产业价值：科技创新、产品质量与客户、供应链管理、信息安全"},
    "S": {"folder": "S-人权与社会", "desc": "人权与社会：员工权益、劳动关系、职业健康与培训、安全生产"},
}

# 子议题前缀 → 二级目录名（兜底显示名）
SUB_PREFIX_FALLBACK = {
    "GA": "GA-可持续发展治理",
    "GB": "GB-规范治理与党建",
    "GC": "GC-商业行为",
    "EA": "EA-应对气候变化",
    "EB": "EB-污染物管控",
    "EC": "EC-资源利用",
    "ED": "ED-生态多样性",
    "DA": "DA-科技创新",
    "DC": "DC-产品质量与客户",
    "DD": "DD-供应链管理",
    "DE": "DE-信息安全与数据",
    "SA": "SA-社会贡献与乡村振兴",
    "SB": "SB-社区关系",
    "SC": "SC-劳工权益",
    "SD": "SD-安全生产",
}

# 跳过的文件名模式（系统文件、隐藏文件等）
SKIP_FILENAME_PATTERNS = [
    re.compile(r"^\."),              # 以 . 开头（含 ._）
    re.compile(r"^~"),               # 以 ~ 开头（Office 临时文件）
    re.compile(r"Thumbs\.db$", re.I),
    re.compile(r"\.DS_Store$", re.I),
    re.compile(r"\.db$", re.I),
]

SKIP_CONTENT_PATTERNS = [
    re.compile(r"定性.*ESG.*资料清单", re.I),
    re.compile(r"定量.*ESG.*资料清单", re.I),
]


# ==============================================================================
# 文本清理与判空
# ==============================================================================

def clean_text(x) -> str:
    """清理文本：去除换行、回车、制表符、空格，返回 strip 后的字符串。"""
    if x is None:
        return ""
    s = str(x).replace("\n", "").replace("\r", "").replace("\t", "").replace(" ", "").strip()
    return s


def is_blank(x) -> bool:
    """判断值是否为空（None / NaN / 空字符串 / 'nan' / 'none'）。"""
    if x is None:
        return True
    if isinstance(x, float) and pd.isna(x):
        return True
    s = str(x).strip()
    return s == "" or s.lower() in {"nan", "none"}


# ==============================================================================
# 编码提取
# ==============================================================================

def extract_code_from_string(text):
    """
    从字符串中提取第一个 ESG 编码（如 GA1、EB10）。
    返回大写编码字符串，或 None。
    """
    if not isinstance(text, str):
        return None
    m = CODE_REGEX.search(text.upper())
    return m.group(1) if m else None


def extract_all_codes_from_string(text):
    """
    从字符串中提取所有 ESG 编码候选（返回列表，大写，保持出现顺序）。
    """
    if not isinstance(text, str):
        return []
    return [m.group(1).upper() for m in CODE_REGEX.finditer(text.upper())]


def parse_prefix_from_text(text):
    """
    从文本提取编码前缀（不带数字）：GA/GB/EA/A 等。
    支持括号格式：应对气候变化（EA）、EA、规范治理(GB)
    """
    if not isinstance(text, str):
        return None
    s = text.upper()

    # 括号中的前缀
    m = re.search(r"[\(（]\s*([AGEDS][A-Z]{0,3})\s*[\)）]", s)
    if m:
        return m.group(1)

    # 纯前缀
    m2 = re.search(r"\b([AGEDS][A-Z]{0,3})\b", s)
    if m2:
        return m2.group(1)

    return None


def parse_serial_number(x):
    """
    从序号/编号字段中抓取第一个整数，用于 prefix + 数字 编码拼接。
    """
    if is_blank(x):
        return None
    s = str(x).strip()
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


# ==============================================================================
# Excel 处理
# ==============================================================================

def unmerge_and_flatten_styles(input_path: str, output_path: str,
                               clear_formats: bool = True):
    """
    1) 取消 Excel 合并单元格并将左上角的值铺满整个合并区域
    2) 可选清理所有样式（字体、填充、边框等），减少格式干扰

    参数:
        input_path:    输入 Excel 文件路径
        output_path:   输出 Excel 文件路径
        clear_formats: 是否清理样式，默认 True
    """
    wb = load_workbook(input_path, data_only=False)

    default_font       = Font()
    default_fill       = PatternFill()
    default_border     = Border()
    default_alignment  = Alignment()
    default_protection = Protection()

    for ws in wb.worksheets:
        # 取消合并并铺值
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            min_row, min_col = mr.min_row, mr.min_col
            max_row, max_col = mr.max_row, mr.max_col
            top_left_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(mr))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(r, c).value = top_left_value

        # 清理样式
        if clear_formats:
            for row in ws.iter_rows():
                for cell in row:
                    cell.font       = default_font
                    cell.fill       = default_fill
                    cell.border     = default_border
                    cell.alignment  = default_alignment
                    cell.protection = default_protection
                    cell.number_format = "General"

    wb.save(output_path)


def find_header_row_for_reference(df_raw, search_rows=200):
    """
    定位参考清单的表头行。

    查找逻辑（优先级递减）：
    1. 同时含「议题」+「指标」+（「序号/编号/编码/类别/维度」之一） → 标准 GEDS sheet
    2. 含「序号」+「资料」→ 总体概况 sheet（特殊格式，无议题/指标列）

    返回表头行索引（0-based），或 None。
    """
    max_scan = min(search_rows, df_raw.shape[0])

    # 标准 sheet：议题 + 指标 + 序号/编码
    for r in range(max_scan):
        row_vals = [clean_text(v) for v in df_raw.iloc[r].tolist()]
        has_topic     = any("议题" in v for v in row_vals if v)
        has_indicator = any("指标" in v for v in row_vals if v)
        has_key = any(
            ("序号" in v) or ("编号" in v) or ("编码" in v) or
            ("识别编码" in v) or ("类别" in v) or ("维度" in v)
            for v in row_vals if v
        )
        if has_topic and has_indicator and has_key:
            return r

    # 总体概况 sheet 兜底：序号 + 资料（必须在不同列中出现）
    for r in range(max_scan):
        row_vals = [clean_text(v) for v in df_raw.iloc[r].tolist()]
        serial_cols = [c for c, v in enumerate(row_vals) if v and "序号" in v]
        list_cols   = [c for c, v in enumerate(row_vals) if v and "资料" in v]
        if serial_cols and list_cols and serial_cols[0] != list_cols[0]:
            return r

    return None


def find_col_idx_by_keywords(df_raw, header_idx, keywords):
    """
    在表头行按关键字匹配列索引。
    keywords 可传单个字符串或列表，返回首个命中的列索引或 None。
    """
    if isinstance(keywords, str):
        keywords = [keywords]
    header_row = [clean_text(v) for v in df_raw.iloc[header_idx].tolist()]
    for c, v in enumerate(header_row):
        for kw in keywords:
            if kw in v:
                return c
    return None


def forward_fill_in_raw(df_raw, start_row, col_indices):
    """
    对指定列从 start_row 开始向下填充空值（处理合并单元格导致的空行）。
    返回填充后的 DataFrame 副本。
    """
    df = df_raw.copy()
    for c in col_indices:
        last = None
        for r in range(start_row, df.shape[0]):
            if c >= df.shape[1]:
                continue
            val = df.iat[r, c]
            if is_blank(val):
                if last is not None:
                    df.iat[r, c] = last
            else:
                last = val
    return df


# ==============================================================================
# 文件过滤
# ==============================================================================

def should_skip_file(filename: str) -> bool:
    """判断文件是否应跳过（系统文件、隐藏文件等）。"""
    for pat in SKIP_FILENAME_PATTERNS:
        if pat.search(filename):
            return True
    return False


def should_skip_content(filename: str) -> bool:
    """判断文件是否为清单文件本身（应跳过不处理）。"""
    for pat in SKIP_CONTENT_PATTERNS:
        if pat.search(filename):
            return True
    return False
