"""
generate_retrieval_queries.py
=============================
ESG 报告框架 → Retrieval Query 自动生成工具（双查询版本）

输入：data/raw/ESG报告框架.xlsx
输出：data/processed/framework_retrieval_queries.json

生成两种查询：
  - retrieval_query: 基础查询（自然语言描述形式）
  - hypothetical_doc: HyDE 假设文档（模拟企业内部文档风格）

阶段：
  1. 解析 Excel，展开合并单元格
  2. 识别叶节点，过滤无效行
  3. 分批调用 LLM 生成基础查询（retrieval_query）
  4. 分批调用 LLM 生成假设文档（hypothetical_doc）
  5. 结果校验与持久化

用法：
  conda run -n esg python3 src/generate_retrieval_queries.py [选项]

选项：
  --model MODEL   LLM 模型名（默认: deepseek-thinking）
  --dry-run       仅解析 Excel + 识别叶节点，不调用 LLM
  --resume        断点续跑，跳过已有结果的叶节点（使用生产模式）
  --debug         调试模式，遇到错误立即停止（默认首次运行开启）

配置文件：
  src/prompts/retrieval_query_base.txt  基础查询 Prompt 模板
  src/prompts/retrieval_query_hyde.txt  HyDE 假设文档 Prompt 模板
"""

import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import threading

from dotenv import load_dotenv
load_dotenv()  # 加载仓库根 .env 文件

import openpyxl
from openai import OpenAI
from progress_tracker import get_tracker

# ── 路径配置 ──────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
INPUT_XLSX = _ROOT / "data" / "raw" / "ESG报告框架.xlsx"
OUTPUT_JSON = _ROOT / "data" / "processed" / "framework_retrieval_queries.json"
PROGRESS_JSON = _ROOT / "data" / "processed" / "_rq_progress.json"  # 断点续跑

# ── Prompt 模板路径 ────────────────────────────────────────────────────────────
PROMPT_DIR = _HERE / "prompts"
PROMPT_BASE_QUERY = PROMPT_DIR / "retrieval_query_base.txt"
PROMPT_HYDE = PROMPT_DIR / "retrieval_query_hyde.txt"

# ── LLM 配置 ──────────────────────────────────────────────────────────────────
LLM_BASE_URL = os.environ.get("LLM_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
LLM_API_KEY = os.environ.get("LLM_API_KEY") or os.environ.get("DASHSCOPE_API_KEY", "")
DEFAULT_MODEL = "deepseek-v3"

# ── LLM 调用参数 ──────────────────────────────────────────────────────────────
LLM_MAX_TOKENS = 8192  # 最大输出 token 数
LLM_TEMPERATURE = 0.3  # 生成温度
LLM_TIMEOUT = float(os.environ.get("LLM_TIMEOUT", "180"))  # 单次请求超时（秒）
LLM_CALL_RETRIES = 3   # 单次 call_llm 内部重试次数（网络/超时）
LLM_RETRY_DELAY = 5    # 重试间隔（秒）

# ── 并发配置 ──────────────────────────────────────────────────────────────────
MAX_CONCURRENT_REQUESTS = 6  # 最大并发数
MAX_BATCH_SIZE = 15  # 单批次最大节点数（防止响应过长被截断）
MAX_RETRIES = 3  # 生产模式下的最大重试次数

# ── 排除清单 ──────────────────────────────────────────────────────────────────
EXCLUDE_L1 = {
    "封面", "目录", "关于本报告", "可持续发展亮点",
    "董事长致辞", "未来展望", "报告附录", "封底",
}


# =============================================================================
# 阶段 1：解析 Excel，展开合并单元格
# =============================================================================

def parse_excel(path: Path) -> list[dict]:
    """
    读取 Excel，展开合并单元格，返回所有数据行（从第 4 行开始）的扁平记录列表。
    每条记录：{row, l1, l2, l3, l4, gloss}
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # 构建合并区域的值映射：(row, col) → 左上角的值
    merge_map: dict[tuple[int, int], str | None] = {}
    for mr in ws.merged_cells.ranges:
        # 左上角的值
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = top_val

    def cell_val(row: int, col: int) -> str:
        """读取单元格值（考虑合并区域），返回 strip 后的字符串，None 返回空串。"""
        if (row, col) in merge_map:
            v = merge_map[(row, col)]
        else:
            v = ws.cell(row=row, column=col).value
        if v is None:
            return ""
        return str(v).strip()

    records = []
    # 第 3 行是表头，数据从第 4 行开始
    for r in range(4, ws.max_row + 1):
        l1 = cell_val(r, 1)   # A列：一级标题
        l2 = cell_val(r, 2)   # B列：二级议题
        l3 = cell_val(r, 3)   # C列：三级指标
        l4 = cell_val(r, 4)   # D列：末级指标
        gloss = cell_val(r, 5)  # E列：释义
        # 跳过完全空行
        if not any([l1, l2, l3, l4, gloss]):
            continue
        records.append({
            "row": r,
            "l1": l1,
            "l2": l2,
            "l3": l3,
            "l4": l4,
            "gloss": gloss,
        })

    print(f"[阶段1] 解析完成，共 {len(records)} 条记录（第4-{ws.max_row}行）")
    return records


# =============================================================================
# 阶段 2：识别叶节点，过滤无效行
# =============================================================================

def _is_empty(s: str) -> bool:
    """判断字段是否为空值（空字符串、"/"）"""
    return s == "" or s == "/"


def identify_leaves(records: list[dict]) -> list[dict]:
    """
    从扁平记录中筛选出叶节点。
    返回叶节点列表，每个包含：
      id, full_path, leaf_title, l1, l2, l3, l4, gloss
    """
    leaves = []
    for rec in records:
        # 规则 1：跳过排除清单
        if rec["l1"] in EXCLUDE_L1:
            continue
        # 规则 2：跳过释义为空或 "/" 的行
        if _is_empty(rec["gloss"]):
            continue
        # 规则 3：确定叶节点标题
        if not _is_empty(rec["l4"]):
            leaf_title = rec["l4"]
        elif not _is_empty(rec["l3"]):
            leaf_title = rec["l3"]
        elif not _is_empty(rec["l2"]):
            # 少数情况：三级指标也为空，二级议题本身即叶节点（如"社会公益"）
            leaf_title = rec["l2"]
        else:
            continue  # 无可用层级

        # 构建完整路径
        parts = [rec["l1"]]
        if not _is_empty(rec["l2"]):
            parts.append(rec["l2"])
        if not _is_empty(rec["l3"]):
            parts.append(rec["l3"])
        if not _is_empty(rec["l4"]):
            parts.append(rec["l4"])
        full_path = " > ".join(parts)

        node_id = f"r{rec['row']:03d}"
        leaves.append({
            "id": node_id,
            "row": rec["row"],
            "full_path": full_path,
            "leaf_title": leaf_title,
            "l1": rec["l1"],
            "l2": rec["l2"],
            "l3": rec["l3"],
            "l4": rec["l4"],
            "gloss": rec["gloss"],
        })

    print(f"[阶段2] 识别出 {len(leaves)} 个叶节点")
    # 按一级标题统计
    from collections import Counter
    l1_counts = Counter(n["l1"] for n in leaves)
    for l1, cnt in l1_counts.items():
        print(f"  · {l1}: {cnt} 个")

    return leaves


# =============================================================================
# 阶段 3：Prompt 加载
# =============================================================================

def load_prompt(prompt_path: Path) -> str:
    """从文件加载 prompt 模板。"""
    if not prompt_path.exists():
        raise FileNotFoundError(f"Prompt 文件不存在: {prompt_path}")
    return prompt_path.read_text(encoding="utf-8").strip()


def build_user_message(
    target_nodes: list[dict],
    context_nodes: list[dict] | None = None,
) -> str:
    """
    将叶节点拼接为 LLM 的 user message。

    Args:
        target_nodes: 待生成的节点（LLM 需要为这些节点输出结果）
        context_nodes: 上下文节点（仅供参考，帮助理解节点间差异，不需要生成）

    Returns:
        格式化的任务数据字符串
    """
    lines = []

    # 如果有上下文节点，先展示完整的兄弟节点列表
    if context_nodes:
        lines.append("## 完整章节列表（供参考，理解各节点间的区分）\n")
        for node in context_nodes:
            is_target = node["id"] in {n["id"] for n in target_nodes}
            marker = "→" if is_target else " "
            lines.append(f"{marker} [{node['id']}] {node['full_path']}")
        lines.append("")
        lines.append("## 待生成节点（只需为以下节点生成输出）\n")

    # 待生成节点的详细信息
    for node in target_nodes:
        line = f"[{node['id']}] 路径：{node['full_path']}\n"
        line += f"    标题：{node['leaf_title']}\n"
        line += f"    释义：{node['gloss']}"
        lines.append(line)

    return "\n\n".join(lines)


# =============================================================================
# 阶段 4：分批调用 LLM 生成
# =============================================================================

def print_progress(message: str, flush: bool = True):
    """打印进度信息，确保立即刷新到终端（兼容后台运行）。"""
    print(message, flush=flush)
    sys.stdout.flush()


def call_llm(client: OpenAI, model: str, user_prompt: str) -> str | None:
    """
    调用 LLM，返回文本内容。失败返回 None。
    内置重试逻辑：网络错误/超时自动重试 LLM_CALL_RETRIES 次。
    """
    for attempt in range(LLM_CALL_RETRIES):
        try:
            resp = client.chat.completions.create(
                model=model,
                max_tokens=LLM_MAX_TOKENS,
                temperature=LLM_TEMPERATURE,
                messages=[
                    {"role": "user", "content": user_prompt},
                ],
            )
            return resp.choices[0].message.content
        except Exception as e:
            if attempt < LLM_CALL_RETRIES - 1:
                delay = LLM_RETRY_DELAY * (attempt + 1)
                print_progress(f"    [LLM] 调用失败: {e}，{delay}s 后重试 ({attempt + 1}/{LLM_CALL_RETRIES})")
                time.sleep(delay)
            else:
                print_progress(f"    [LLM] 调用失败（已重试 {LLM_CALL_RETRIES} 次）: {e}")
                return None


def parse_json_response(text: str) -> list[dict] | None:
    """
    从 LLM 返回文本中解析 JSON 数组。
    仅处理 markdown 代码块包裹，其他格式问题应通过 prompt 约束解决。
    """
    if text is None:
        return None

    # 去除 markdown 代码块
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    # 直接解析
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    return None


def process_batch(
    client: OpenAI,
    model: str,
    batch: list[dict],
    batch_name: str,
    instruction_prompt: str,
    field_name: str,
    raw_cache: dict[str, str] | None = None,
    debug_mode: bool = True,
    context_nodes: list[dict] | None = None,
    progress_dir: Path | None = None,
) -> tuple[dict[str, str], list[dict]]:
    """
    处理一个批次：调用 LLM，解析返回，校验 id 匹配。

    Args:
        batch: 待生成的节点列表
        instruction_prompt: 指令提示词（不含具体任务数据）
        debug_mode: 调试模式（True=遇错即停，False=重试后跳过）
        context_nodes: 完整的兄弟节点列表（作为上下文参考）

    Returns:
        (成功字典 {id: value}, 失败节点列表)
    """
    expected_ids = {n["id"] for n in batch}

    # 构建完整的 user prompt：指令 + 任务数据（含上下文）
    task_data = build_user_message(batch, context_nodes)
    full_user_prompt = f"{instruction_prompt}\n\n# 任务数据\n\n{task_data}"

    max_attempts = 1 if debug_mode else MAX_RETRIES
    last_raw = None

    for attempt in range(max_attempts):
        if attempt == 0:
            print_progress(f"  → 发送批次「{batch_name}」({len(batch)} 条)")
        else:
            print_progress(f"  ↻ 重试批次「{batch_name}」(第 {attempt + 1}/{max_attempts} 次)")

        raw = call_llm(client, model, full_user_prompt)

        if raw is None:
            print_progress(f"  ✗ 批次「{batch_name}」LLM 调用失败")
            continue

        last_raw = raw

        # 保存原始返回到缓存
        if raw_cache is not None:
            raw_cache[batch_name] = raw

        results = parse_json_response(raw)

        if results is None:
            print_progress(f"  ✗ 批次「{batch_name}」JSON 解析失败")
            continue

        # 校验解析结果
        success = {}
        found_ids = set()
        for item in results:
            nid = item.get("id", "")
            value = item.get(field_name, "")
            if nid in expected_ids and value:
                success[nid] = value
                found_ids.add(nid)

        missing = expected_ids - found_ids
        failed = [n for n in batch if n["id"] in missing]

        if failed:
            print_progress(f"  ⚠ 批次「{batch_name}」成功 {len(success)}/{len(batch)}，"
                           f"缺失 {len(failed)} 条: {[n['id'] for n in failed]}")
        else:
            print_progress(f"  ✓ 批次「{batch_name}」全部成功 ({len(success)} 条)")

        return success, failed

    # 所有重试都失败
    if debug_mode:
        # 调试模式：保存原始返回到文件并抛出异常
        debug_file = (progress_dir or PROGRESS_JSON.parent) / f"_debug_parse_error_{batch_name}.txt"
        with open(debug_file, "w", encoding="utf-8") as f:
            f.write(f"批次名: {batch_name}\n")
            f.write(f"节点数: {len(batch)}\n")
            f.write(f"节点ID: {[n['id'] for n in batch]}\n")
            f.write(f"{'='*60}\n")
            f.write("LLM 原始返回:\n")
            f.write(f"{'='*60}\n")
            f.write(last_raw or "(无返回)")
        print_progress(f"  💾 原始返回已保存到: {debug_file}")
        raise RuntimeError(f"批次「{batch_name}」处理失败，已停止执行。请检查: {debug_file}")
    else:
        # 生产模式：跳过失败批次
        print_progress(f"  ✗ 批次「{batch_name}」重试 {max_attempts} 次后仍失败，跳过")
        return {}, batch


def run_generation_phase(
    client: OpenAI,
    model: str,
    leaves: list[dict],
    result_map: dict[str, str],
    instruction_prompt: str,
    field_name: str,
    phase_name: str,
    save_progress_callback=None,
    debug_mode: bool = True,
    progress_dir: Path | None = None,
) -> dict[str, str]:
    """
    运行一个生成阶段（基础查询或假设文档），支持并发处理。

    Args:
        instruction_prompt: 指令提示词（将作为 user message 的开头）
        debug_mode: 调试模式，遇到解析失败立即停止

    Returns:
        更新后的 result_map
    """
    from collections import OrderedDict
    import math

    result_lock = threading.Lock()

    # 按一级标题分组
    l1_groups: OrderedDict[str, list[dict]] = OrderedDict()
    for n in leaves:
        l1 = n["l1"]
        if l1 not in l1_groups:
            l1_groups[l1] = []
        l1_groups[l1].append(n)

    # 构建批次：(批次名, 待生成节点, 完整上下文节点)
    # 大组均分，保留完整兄弟节点作为上下文
    final_batches: list[tuple[str, list[dict], list[dict] | None]] = []
    for l1_name, group_nodes in l1_groups.items():
        if len(group_nodes) <= MAX_BATCH_SIZE:
            # 小批次：无需拆分，无需额外上下文
            final_batches.append((l1_name, group_nodes, None))
        else:
            # 大批次：均分，并提供完整的兄弟节点作为上下文
            n_splits = math.ceil(len(group_nodes) / MAX_BATCH_SIZE)
            split_size = math.ceil(len(group_nodes) / n_splits)  # 均分
            for i in range(n_splits):
                start = i * split_size
                end = min(start + split_size, len(group_nodes))
                sub_nodes = group_nodes[start:end]
                sub_name = f"{l1_name}_{i + 1}"
                # 传递完整的组节点作为上下文
                final_batches.append((sub_name, sub_nodes, group_nodes))

    total_batches = len(final_batches)
    total_nodes = len(leaves)
    completed_nodes = len([n for n in leaves if n["id"] in result_map])

    print_progress(f"\n[{phase_name}] 开始处理 {total_nodes} 个节点，分 {total_batches} 批...")
    print_progress(f"  已完成: {completed_nodes}/{total_nodes}")
    print_progress(f"  并发数: {MAX_CONCURRENT_REQUESTS}")
    print_progress(f"  调试模式: {'ON (遇到解析失败会停止)' if debug_mode else 'OFF'}")

    # 准备待处理的批次
    pending_batches = []
    for batch_name, batch_nodes, context_nodes in final_batches:
        pending = [n for n in batch_nodes if n["id"] not in result_map]
        if pending:
            pending_batches.append((batch_name, pending, context_nodes))
        else:
            print_progress(f"  ⏭「{batch_name}」全部已完成，跳过")

    if not pending_batches:
        print_progress(f"  所有批次已完成，跳过")
        return result_map, []

    print_progress(f"  待处理批次: {len(pending_batches)}")

    # 并发处理函数
    all_failed: list[dict] = []

    def process_one_batch(batch_info):
        batch_name, batch_nodes, context_nodes = batch_info
        success, failed = process_batch(
            client, model, batch_nodes, batch_name, instruction_prompt, field_name,
            None, debug_mode, context_nodes, progress_dir
        )
        return batch_name, success, failed

    # 使用线程池并发处理
    with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_REQUESTS) as executor:
        futures = {executor.submit(process_one_batch, batch): batch for batch in pending_batches}

        for future in as_completed(futures):
            batch_name, success, failed = future.result()

            with result_lock:
                result_map.update(success)
                all_failed.extend(failed)
                completed_nodes = len([n for n in leaves if n["id"] in result_map])
                print_progress(f"  📊 进度: {completed_nodes}/{total_nodes} 节点")

                # 每批完成后保存进度
                if save_progress_callback and success:
                    save_progress_callback()

    return result_map, all_failed


# =============================================================================
# 阶段 5-6：进度管理与主流程
# =============================================================================

def load_progress(progress_path: Path | None = None) -> dict:
    """
    加载已有进度。
    返回格式: {"retrieval_query": {id: value}, "hypothetical_doc": {id: value}}
    """
    pj = progress_path or PROGRESS_JSON
    if pj.exists():
        with open(pj, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data
    return {"retrieval_query": {}, "hypothetical_doc": {}}


def save_progress(progress: dict, progress_path: Path | None = None):
    """保存进度到磁盘。"""
    pj = progress_path or PROGRESS_JSON
    pj.parent.mkdir(parents=True, exist_ok=True)
    with open(pj, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def cleanup_temp_files(progress_path: Path | None = None):
    """清理临时文件（进度文件、调试文件、原始缓存）。"""
    pj = progress_path or PROGRESS_JSON
    patterns = [
        pj,
        *pj.parent.glob("_debug_parse_error_*.txt"),
        *pj.parent.glob("_rq_raw_cache_*.json"),
    ]
    cleaned = []
    for p in patterns:
        if isinstance(p, Path) and p.exists():
            p.unlink()
            cleaned.append(p.name)
    return cleaned


def main():
    parser = argparse.ArgumentParser(description="ESG 报告框架 Retrieval Query 生成（双查询版本）")
    parser.add_argument("--model", default=DEFAULT_MODEL,
                        help=f"LLM 模型名 (默认: {DEFAULT_MODEL})")
    parser.add_argument("--dry-run", action="store_true",
                        help="仅解析 Excel + 识别叶节点，不调用 LLM")
    parser.add_argument("--resume", action="store_true",
                        help="断点续跑，跳过已有结果的叶节点")
    parser.add_argument("--debug", action="store_true",
                        help="调试模式：遇到错误立即停止（默认关闭）")
    parser.add_argument("--retry-failed", action="store_true",
                        help="重跑失败节点：读取已有输出文件，筛出 status=='needs_manual_review' 的节点重新生成")
    parser.add_argument("--tracker", type=str, default=None,
                        help="Web UI 进度追踪 run_id（内部使用）")
    parser.add_argument(
        "--project-dir",
        type=str,
        default=None,
        metavar="DIR",
        help=(
            "企业项目目录（如 projects/艾森股份_2025）。"
            "不传则使用旧的 data/ 路径（向后兼容）。"
        ),
    )
    args = parser.parse_args()

    # ── 路径初始化（多企业支持，向后兼容）────────────────────────────────────
    from config import get_paths
    paths = get_paths(args.project_dir)
    paths.processed_dir.mkdir(parents=True, exist_ok=True)

    input_xlsx    = paths.framework_xlsx
    output_json   = paths.framework_queries
    progress_json = paths.rq_progress
    progress_dir  = paths.processed_dir

    # Web UI 进度追踪
    tracker = get_tracker(args, "generate_retrieval_queries")

    # 耗时统计
    from stage_timer import StageTimer
    timer = StageTimer()

    # 只有明确指定 --debug 才进入调试模式，默认使用生产模式（带重试）
    debug_mode = args.debug

    print_progress(f"{'='*60}")
    print_progress(f"ESG 报告框架 Retrieval Query 生成（双查询版本）")
    print_progress(f"模型: {args.model}")
    print_progress(f"输入: {input_xlsx}")
    print_progress(f"输出: {output_json}")
    print_progress(f"并发数: {MAX_CONCURRENT_REQUESTS}")
    print_progress(f"模式: {'调试模式（遇错即停）' if debug_mode else f'生产模式（失败重试 {MAX_RETRIES} 次）'}")
    print_progress(f"{'='*60}\n")

    # ── 阶段 1 ──
    timer.start("阶段 1：解析 Excel")
    tracker.set_stage("Parse Excel")
    records = parse_excel(input_xlsx)

    # ── 阶段 2 ──
    timer.start("阶段 2：识别叶节点")
    tracker.set_stage("Identify leaves")
    leaves = identify_leaves(records)

    if not leaves:
        print_progress("未找到任何叶节点，退出。")
        return

    if args.dry_run:
        print_progress("\n[dry-run] 叶节点列表：")
        for n in leaves:
            print_progress(f"  {n['id']} | {n['full_path']}")
            print_progress(f"       释义: {n['gloss'][:80]}...")
        print_progress(f"\n[dry-run] 共 {len(leaves)} 个叶节点，退出。")
        return

    # ── 初始化 LLM 客户端 ──
    client = OpenAI(
        base_url=LLM_BASE_URL,
        api_key=LLM_API_KEY,
        timeout=LLM_TIMEOUT,
    )

    # ── 加载已有进度 ──
    progress = {"retrieval_query": {}, "hypothetical_doc": {}}
    if args.resume:
        progress = load_progress(progress_json)
        print_progress(f"  已加载历史进度: "
                       f"retrieval_query={len(progress['retrieval_query'])} 条, "
                       f"hypothetical_doc={len(progress['hypothetical_doc'])} 条")

    # 创建保存进度的回调函数
    def save_progress_callback():
        save_progress(progress, progress_json)

    # ── --retry-failed：替换 leaves 为失败节点，复用现有流程 ──
    if args.retry_failed:
        if not output_json.exists():
            print_progress(f"[错误] 输出文件不存在，无法重跑失败节点：{output_json}")
            sys.exit(1)
        with open(output_json, "r", encoding="utf-8") as f:
            existing_output = json.load(f)
        failed_leaf_ids = {
            e["id"] for e in existing_output if e.get("status") == "needs_manual_review"
        }
        if not failed_leaf_ids:
            print_progress("没有需要重跑的失败节点（status=='needs_manual_review'），退出")
            sys.exit(0)
        leaves = [n for n in leaves if n["id"] in failed_leaf_ids]
        print_progress(f"  --retry-failed：筛出 {len(leaves)} 个失败节点重跑")
        # 加载已有进度，清除失败节点的旧 null 值
        progress = load_progress(progress_json)
        for leaf_id in failed_leaf_ids:
            progress["retrieval_query"].pop(leaf_id, None)
            progress["hypothetical_doc"].pop(leaf_id, None)

    # ── 加载 Prompt 模板 ──
    prompt_base_query = load_prompt(PROMPT_BASE_QUERY)
    prompt_hyde = load_prompt(PROMPT_HYDE)

    # ── 阶段 3：生成基础查询 ──
    timer.start("阶段 3：生成基础查询")
    print_progress(f"\n{'='*60}")
    print_progress("阶段 3: 生成基础查询 (retrieval_query)")
    print_progress(f"{'='*60}")
    tracker.set_stage("Base queries", total=len(leaves))

    progress["retrieval_query"], failed_rq = run_generation_phase(
        client=client,
        model=args.model,
        leaves=leaves,
        result_map=progress["retrieval_query"],
        instruction_prompt=prompt_base_query,
        field_name="retrieval_query",
        phase_name="基础查询",
        save_progress_callback=save_progress_callback,
        debug_mode=debug_mode,
        progress_dir=progress_dir,
    )
    save_progress(progress, progress_json)

    # ── 阶段 4：生成假设文档 ──
    timer.start("阶段 4：生成假设文档")
    print_progress(f"\n{'='*60}")
    print_progress("阶段 4: 生成假设文档 (hypothetical_doc)")
    print_progress(f"{'='*60}")
    tracker.set_stage("HyDE docs", total=len(leaves))

    progress["hypothetical_doc"], failed_hd = run_generation_phase(
        client=client,
        model=args.model,
        leaves=leaves,
        result_map=progress["hypothetical_doc"],
        instruction_prompt=prompt_hyde,
        field_name="hypothetical_doc",
        phase_name="假设文档",
        save_progress_callback=save_progress_callback,
        debug_mode=debug_mode,
        progress_dir=progress_dir,
    )
    save_progress(progress, progress_json)

    # ── 输出最终 JSON ──
    timer.start("输出 JSON")
    print_progress(f"\n{'='*60}")
    print_progress("[输出] 生成最终 JSON...")
    tracker.set_stage("Validate")

    output = []
    for n in leaves:
        entry = {
            "id": n["id"],
            "row": n["row"],
            "full_path": n["full_path"],
            "leaf_title": n["leaf_title"],
            "l1": n["l1"],
            "l2": n["l2"],
            "l3": n["l3"],
            "l4": n["l4"],
            "gloss": n["gloss"],
        }

        # 基础查询
        if n["id"] in progress["retrieval_query"]:
            entry["retrieval_query"] = progress["retrieval_query"][n["id"]]
        else:
            entry["retrieval_query"] = None

        # 假设文档
        if n["id"] in progress["hypothetical_doc"]:
            entry["hypothetical_doc"] = progress["hypothetical_doc"][n["id"]]
        else:
            entry["hypothetical_doc"] = None

        # 标记需要人工审查的条目
        if entry["retrieval_query"] is None or entry["hypothetical_doc"] is None:
            entry["status"] = "needs_manual_review"

        output.append(entry)

    output_json.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # 统计
    total = len(output)
    rq_ok = sum(1 for e in output if e.get("retrieval_query"))
    hd_ok = sum(1 for e in output if e.get("hypothetical_doc"))
    all_ok = sum(1 for e in output if e.get("retrieval_query") and e.get("hypothetical_doc"))
    needs_review = total - all_ok

    print_progress(f"\n{'='*60}")
    print_progress(f"完成！共 {total} 个叶节点")
    print_progress(f"  ✓ retrieval_query 成功: {rq_ok}/{total}")
    print_progress(f"  ✓ hypothetical_doc 成功: {hd_ok}/{total}")
    print_progress(f"  ✓ 全部完成: {all_ok}/{total}")
    print_progress(f"  ⚠ 待人工审查: {needs_review}")
    if needs_review > 0:
        all_failed_ids = [n["id"] for n in failed_rq] + [n["id"] for n in failed_hd]
        unique_failed = list(dict.fromkeys(all_failed_ids))
        if unique_failed:
            print_progress(f"  失败节点 ID（{len(unique_failed)} 个）：{unique_failed}")
    print_progress(f"输出文件: {output_json}")
    print_progress(f"{'='*60}")

    # 全部成功后清理临时文件
    if needs_review == 0:
        cleaned = cleanup_temp_files(progress_json)
        if cleaned:
            print_progress(f"  🧹 已清理临时文件: {', '.join(cleaned)}")

    timer.report()
    if needs_review > 0:
        tracker.set_partial_failed(needs_review, unique_failed)
    tracker.complete()
    if needs_review > 0:
        sys.exit(1)
if __name__ == "__main__":
    main()
