"""
align_evidence.py
=================
ESG 双证据融合对齐脚本：将路径编码（先验）与语义向量检索（后验）交叉验证，
输出三类状态的对齐表：
  ✅ 一致：路径编码在语义 Top-K 内，直接采用
  ⚠️  疑似错位：路径有编码但语义不匹配，需人工确认
  🔍 无路径标签但语义命中：兜底文件夹文件，自动建议归属

开发进度：
  阶段一（完成）：加载清单映射 + 扫描文件
  阶段二（完成）：文本提取与分块 → src/extractors.py
  阶段三（完成）：embedding 构建 → build_indicator_queries / compute_embeddings /
                                    build_indicator_collection / embed_chunks /
                                    save_emb_cache / load_emb_cache
  阶段四（完成）：语义检索 + 一致性判断 → semantic_search_batch / classify_consistency /
                                           align_chunks / print_phase4_summary
  阶段五（完成）：输出对齐表 → write_alignment_excel / print_phase5_summary

阶段二缓存机制：
  - 阶段 2a 文本提取结果写入 sections_cache.json（JSON 文件），下次运行直接加载，跳过重提取
  - 阶段 2b 分块结果写入 chunks_cache.json（JSON 文件），下次运行直接加载，跳过重分块
  - 阶段三 chunk embedding 写入 chunks_emb_cache.npz（numpy .npz），下次运行直接加载，跳过重算
  - 使用 --rebuild embedding 重算向量（最常用，如换模型或改维度后）
  - 使用 --rebuild chunk 重新分块（修改分块参数后）
  - 使用 --rebuild extract 全部重新提取（资料文件夹有变化时）
  - 使用 --rebuild all 等同 extract

运行方式：
    # 向后兼容（使用旧的 data/ 路径）
    python3 src/align_evidence.py

    # 多企业模式（路径隔离到 projects/ 子目录）
    python3 src/align_evidence.py --project-dir projects/艾森股份_2025
"""

import os
import sys
import time
import argparse
import numpy as np
from tqdm import tqdm

from progress_tracker import get_tracker

# 将 src/ 加入 sys.path，以便 import 同目录的脚本模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from esg_utils import (
    CODE_REGEX,
    DIMENSION_META,
    should_skip_file,
    should_skip_content,
    extract_all_codes_from_string,
)
from data_list_v2 import load_esg_mapping_from_reference_excel, find_best_code_in_path
from generate_folder_structure import load_full_esg_info
from extractors import (
    configure_vlm_context,
    load_vlm_cache,
    save_vlm_cache,
    extract_pdf,
    extract_docx,
    extract_doc,
    extract_xlsx,
    extract_xls,
    extract_pptx,
    extract_ppt,
    extract_image,
    extract_sections,
    make_chunks_from_sections,
    chunk_params,
    get_text_for_embedding,  # Phase 2：Embedding 字段选择
)

from config import (
    OPENAI_API_KEY,
    OPENAI_BASE_URL,
    EMBEDDING_MODEL,
    EMBEDDING_INSTRUCT,
    EMBEDDING_TOP_K,
    CONSISTENCY_TOPN,
    EXTRA_RELEVANCE_THRESHOLD,
    MIN_RELEVANCE_SCORE,
    ENABLE_TABLE_SUMMARY,  # Phase 2：表格摘要开关
    VLM_MODEL,
    get_paths,
)


# ==============================================================================
# 用户配置：只改这里
# ==============================================================================
_HERE = os.path.dirname(os.path.abspath(__file__))  # src/ 目录
_ROOT = os.path.dirname(_HERE)                       # 项目根目录
# 算法参数、服务地址、模型名统一在 src/config.py 中管理
# 企业相关路径通过 --project-dir 参数动态决定（见 parse_args + get_paths）
# ==============================================================================


# ==============================================================================
# 命令行参数解析
# ==============================================================================

def parse_args():
    """
    解析命令行参数。

    --project-dir  多企业并行（不传则向后兼容，使用旧的 data/ 路径）
    --rebuild      缓存重建级别（extract / chunk / embedding / all）
    --tracker      Web UI Pipeline 调用（CLI 直接运行时不传）
    """
    parser = argparse.ArgumentParser(
        description="ESG 双证据对齐流水线",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--project-dir",
        type=str,
        default=None,
        metavar="DIR",
        help=(
            "企业项目目录（如 projects/艾森股份_2025）。"
            "不传则使用旧的 data/ 路径（向后兼容）。"
        ),
    )
    parser.add_argument(
        "--rebuild",
        type=str,
        default=None,
        choices=["extract", "chunk", "embedding", "all"],
        metavar="LEVEL",
        help=(
            "重建缓存：extract=全部重新提取，chunk=重新分块，"
            "embedding=重算向量（最常用），all=等同extract。"
            "不传则复用已有缓存。"
        ),
    )
    parser.add_argument(
        "--tracker",
        type=str,
        default=None,
        metavar="RUN_ID",
        help="Web UI 进度追踪 run_id（内部使用，CLI 直接运行时不传）",
    )
    return parser.parse_args()


# ==============================================================================
# 缓存重建 — _cleanup_caches
# ==============================================================================

def _cleanup_caches(paths, level: str):
    """
    根据重建级别删除对应缓存文件。

    级联关系：extract ⊃ chunk ⊃ embedding
      - extract:   sections + vlm + chunks + table_summary + emb + chroma
      - chunk:     chunks + table_summary + emb + chroma
      - embedding: emb + chroma

    Args:
        paths: ProjectPaths 实例
        level: "extract" | "chunk" | "embedding"
    """
    import shutil
    from pathlib import Path

    def _rm(p):
        """安全删除文件，不存在时静默跳过。"""
        p = Path(p) if not isinstance(p, Path) else p
        if p.is_file():
            p.unlink()
            print(f"    删除: {p.name}")

    def _rmdir(p):
        """安全删除目录，不存在时静默跳过。"""
        p = Path(p) if not isinstance(p, Path) else p
        if p.is_dir():
            shutil.rmtree(p)
            print(f"    删除: {p.name}/")

    print(f"  [缓存重建] 级别: {level}")

    # embedding 级别：emb + chroma
    if level in ("extract", "chunk", "embedding"):
        _rm(paths.emb_cache)
        _rmdir(paths.chroma_dir)

    # chunk 级别：chunks + table_summary（+ 上面的 emb/chroma）
    if level in ("extract", "chunk"):
        _rm(paths.chunk_cache)
        _rm(paths.table_summary_cache)

    # extract 级别：sections + vlm（+ 上面所有）
    if level == "extract":
        _rm(paths.section_cache)
        _rm(paths.vlm_cache)


# ==============================================================================
# 阶段一 — 函数一：加载指标详情映射
# ==============================================================================

def load_indicator_details(reference_excel_path: str) -> dict:
    """
    从参考清单构建完整指标映射，包含 requirement 字段。

    返回：
        {
            code: {
                "topic":       str,  # 议题
                "indicator":   str,  # 指标
                "requirement": str,  # 资料需求描述（阶段三用于拼接 embedding 查询文本）
            },
            ...
        }
    共 189 条（对应清单中全部定性指标）。
    """
    full_info_list = load_full_esg_info(reference_excel_path)

    details = {}
    for item in full_info_list:
        code = item["code"]
        details[code] = {
            "topic":       item.get("topic", ""),
            "indicator":   item.get("indicator", ""),
            "requirement": item.get("requirement", ""),
        }
    return details


# ==============================================================================
# 阶段一 — 函数二：扫描目标文件夹，生成文件记录列表
# ==============================================================================

def scan_target_files(target_folder: str, esg_mapping: dict) -> list:
    """
    遍历 target_folder，为每个有效文件生成一条 FileRecord。

    FileRecord 字段：
        file_path      - 绝对路径
        relative_path  - 相对 target_folder 的路径
        file_name      - 文件名（含扩展名）
        folder_code    - 从路径提取的最优 ESG 编码（兜底文件夹为 None）
        extension      - 小写扩展名（如 ".pdf"、".docx"）

    过滤规则：
        - should_skip_file()：系统文件、隐藏文件（. 或 ~ 开头）、Thumbs.db 等
        - should_skip_content()：ESG 资料清单文件本身
        - 📋说明.txt：文件夹模板遗留的说明文件，跳过
    """
    file_records = []

    for current_root, dirs, files in os.walk(target_folder):
        dirs.sort()
        files.sort()

        for filename in files:
            if should_skip_file(filename):
                continue
            if should_skip_content(filename):
                continue
            if filename == "📋说明.txt":
                continue

            abs_path   = os.path.join(current_root, filename)
            rel_path   = os.path.relpath(abs_path, target_folder)
            path_parts = rel_path.split(os.sep)

            best_code, _part, _idx, _in_map = find_best_code_in_path(path_parts, esg_mapping)
            _, ext = os.path.splitext(filename)

            file_records.append({
                "file_path":     abs_path,
                "relative_path": rel_path,
                "file_name":     filename,
                "folder_code":   best_code,
                "extension":     ext.lower(),
            })

    return file_records


# ==============================================================================
# 阶段一 — 打印摘要
# ==============================================================================

def print_phase1_summary(esg_mapping: dict, indicator_details: dict,
                         file_records: list) -> None:
    """打印阶段一的统计摘要到控制台。"""
    req_nonempty = sum(
        1 for v in indicator_details.values() if v.get("requirement", "").strip()
    )

    ext_counter: dict = {}
    for rec in file_records:
        raw_ext = rec["extension"]
        display = raw_ext.lstrip(".").upper() if raw_ext else "无后缀"
        ext_counter[display] = ext_counter.get(display, 0) + 1

    sorted_exts = sorted(ext_counter.items(), key=lambda x: -x[1])
    top_exts    = sorted_exts[:6]
    others      = sum(cnt for _, cnt in sorted_exts[6:])

    ext_parts = [f"{name}: {cnt}" for name, cnt in top_exts]
    if others > 0:
        ext_parts.append(f"其他: {others}")

    fallback_count = sum(1 for rec in file_records if rec["folder_code"] is None)

    print(f"  ✓ 加载 {len(indicator_details)} 个指标编码"
          f"（含 requirement 字段：{req_nonempty} 条非空）")
    print(f"  ✓ 扫描到 {len(file_records)} 个文件")
    print(f"    {', '.join(ext_parts)}")
    print(f"  ✓ 兜底文件夹文件（folder_code=None）: {fallback_count} 个")


# ==============================================================================
# 阶段二缓存 — 读写 chunk_records（JSON 持久化）
# ==============================================================================

def save_chunks_cache(chunks_data: dict, cache_path: str) -> None:
    """
    将 chunks 数据序列化为 JSON 写入 cache_path。

    Phase 1 表格优化新结构：
        {
            "parents": {parent_id: parent_text, ...},
            "chunks": [chunk_dict, ...]
        }

    文件不存在时自动创建；写入失败时打印警告，不中断主流程。
    """
    import json
    try:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(chunks_data, f, ensure_ascii=False, indent=2)
        size_mb = os.path.getsize(cache_path) / 1024 / 1024
        chunk_count = len(chunks_data.get("chunks", []))
        parent_count = len(chunks_data.get("parents", {}))
        print(f"  ✓ chunk 缓存已写入：{cache_path}（{size_mb:.1f} MB）")
        print(f"    chunks: {chunk_count} 条, parents: {parent_count} 条")
    except Exception as e:
        print(f"  [警告] chunk 缓存写入失败：{e}")


def load_chunks_cache(cache_path: str) -> dict | None:
    """
    从 cache_path 读取已缓存的 chunks 数据。

    返回：
        {"parents": {...}, "chunks": [...]}

    文件不存在或解析失败时返回 None（触发重新提取）。
    """
    import json
    if not os.path.exists(cache_path):
        return None
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and "chunks" in data:
            return data
        print(f"  [警告] 缓存文件格式异常，将重新提取")
        return None
    except Exception as e:
        print(f"  [警告] 缓存文件读取失败：{e}，将重新提取")
        return None


# ==============================================================================
# 阶段 2a 缓存 — 读写 sections（JSON 持久化，按 relative_path 分组）
# ==============================================================================

def save_sections_cache(all_sections: dict, cache_path: str) -> None:
    """
    将 {relative_path: [section_dicts]} 序列化为 JSON 写入 cache_path。
    文件不存在时自动创建；写入失败时打印警告，不中断主流程。
    """
    import json
    try:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(all_sections, f, ensure_ascii=False, indent=2)
        size_mb = os.path.getsize(cache_path) / 1024 / 1024
        total_secs = sum(len(v) for v in all_sections.values())
        print(f"  ✓ sections 缓存已写入：{cache_path}（{size_mb:.1f} MB，"
              f"{len(all_sections)} 个文件，{total_secs} 个 sections）")
    except Exception as e:
        print(f"  [警告] sections 缓存写入失败：{e}")


def load_sections_cache(cache_path: str) -> dict | None:
    """
    从 cache_path 读取已缓存的 sections。
    返回 {relative_path: [section_dicts]} 或 None（触发重新提取）。
    """
    import json
    if not os.path.exists(cache_path):
        return None
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            print(f"  [警告] sections 缓存格式异常（非 dict），将重新提取")
            return None
        return data
    except Exception as e:
        print(f"  [警告] sections 缓存读取失败：{e}，将重新提取")
        return None


# ==============================================================================
# 阶段三缓存 — 读写 chunk embedding（numpy .npz 持久化）
# ==============================================================================

def save_emb_cache(chunk_records_with_emb: list, cache_path: str) -> None:
    """
    将 chunk embedding 向量矩阵以 numpy .npz 压缩格式写入 cache_path。

    存储内容：
        embeddings - float32 矩阵 (N, dim)，N = chunk 数量，dim = 向量维度
        valid_mask - bool 数组 (N,)，标记每条 embedding 是否有效

    embedding 为 None 的条目（空文本 chunk）→ 全零向量 + valid_mask[i] = False。
    写入失败时打印警告，不中断主流程。
    """
    try:
        n = len(chunk_records_with_emb)
        if n == 0:
            print("  [警告] embedding 缓存：无记录，跳过写入")
            return

        # 从第一条有效 embedding 推断维度
        dim = 0
        for rec in chunk_records_with_emb:
            emb = rec.get("embedding")
            if emb is not None and len(emb) > 0:
                dim = len(emb)
                break
        if dim == 0:
            print("  [警告] embedding 缓存：无有效 embedding，跳过写入")
            return

        matrix = np.zeros((n, dim), dtype=np.float32)
        mask = np.zeros(n, dtype=bool)

        for i, rec in enumerate(chunk_records_with_emb):
            emb = rec.get("embedding")
            if emb is not None and len(emb) > 0:
                matrix[i] = emb
                mask[i] = True

        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        np.savez_compressed(cache_path, embeddings=matrix, valid_mask=mask)

        size_mb = os.path.getsize(cache_path) / 1024 / 1024
        valid_count = int(mask.sum())
        print(f"  ✓ embedding 缓存已写入：{cache_path}")
        print(f"    （{size_mb:.1f} MB，{n} 条，{valid_count} 有效）")
    except Exception as e:
        print(f"  [警告] embedding 缓存写入失败：{e}")


def load_emb_cache(cache_path: str, expected_count: int) -> list | None:
    """
    从 cache_path 读取已缓存的 chunk embedding。

    参数：
        cache_path      - .npz 文件路径
        expected_count  - 期望的 chunk 数量（与 chunk_records 长度一致）

    返回：
        List[list | None]，长度 == expected_count。
        valid_mask=False 的位置返回 None，valid_mask=True 的位置返回 list[float]。
        文件不存在 / 数量不匹配 / 读取异常 → 返回 None（触发重新计算）。
    """
    if not os.path.exists(cache_path):
        return None
    try:
        data = np.load(cache_path)
        embeddings = data["embeddings"]  # (N, dim) float32
        valid_mask = data["valid_mask"]  # (N,) bool

        if embeddings.shape[0] != expected_count:
            print(f"  [提示] embedding 缓存条数不一致"
                  f"（缓存 {embeddings.shape[0]} vs 当前 {expected_count}），将重新计算")
            return None

        result = []
        for i in range(expected_count):
            if valid_mask[i]:
                result.append(embeddings[i].tolist())
            else:
                result.append(None)
        return result
    except Exception as e:
        print(f"  [警告] embedding 缓存读取失败：{e}，将重新计算")
        return None


# ==============================================================================
# 打印脚本标题
# ==============================================================================

def print_header(company_name: str) -> None:
    """打印脚本启动标题横幅。"""
    print("═" * 39)
    print(f"  ESG 双证据融合对齐 — {company_name}")
    print("═" * 39)
    print()


def print_config_summary(paths, company_name: str, rebuild: str = None) -> None:
    """
    打印本次运行的关键配置项，帮助开发者在日志开头快速核查配置是否正确。
    只打印"容易改错或直接影响本次运行路径"的配置，算法阈值等不打印。
    """
    section_cache_path = str(paths.section_cache)
    chunk_cache_path = str(paths.chunk_cache)
    emb_cache_path = str(paths.emb_cache)

    sec_cache_exists = os.path.exists(section_cache_path)
    sec_cache_status = (f"存在（{os.path.getsize(section_cache_path)/1024/1024:.1f} MB）"
                        if sec_cache_exists else "不存在（将重新提取）")

    cache_exists = os.path.exists(chunk_cache_path)
    cache_status = (f"存在（{os.path.getsize(chunk_cache_path)/1024/1024:.1f} MB）"
                    if cache_exists else "不存在（将重新分块）")

    emb_cache_exists = os.path.exists(emb_cache_path)
    emb_cache_status = (f"存在（{os.path.getsize(emb_cache_path)/1024/1024:.1f} MB）"
                        if emb_cache_exists else "不存在（将重新计算）")

    print("[配置摘要]")
    print(f"  公司名称        : {company_name}")
    print(f"  资料文件夹      : {paths.materials_dir}")
    print(f"  参考清单        : {paths.checklist_xlsx}")
    print(f"  输出目录        : {paths.processed_dir}")
    print(f"  {'─' * 37}")
    if rebuild:
        print(f"  缓存重建        : ⚠️  --rebuild {rebuild}")
    else:
        print(f"  缓存重建        : 否（复用已有缓存）")
    print(f"  Sections 缓存   : {sec_cache_status}")
    print(f"  分块缓存状态    : {cache_status}")
    print(f"  Embedding 缓存状态: {emb_cache_status}")
    print(f"  {'─' * 37}")
    print(f"  Embedding 服务  : DashScope SDK（{EMBEDDING_MODEL}）")
    print(f"  VLM 模型        : DashScope {VLM_MODEL}")
    print(f"  ChromaDB 路径   : {paths.chroma_dir}")
    print()


# ==============================================================================
# 阶段三 — 函数一：构建指标查询文本
# ==============================================================================

def build_indicator_queries(indicator_details: dict, enhanced_query_path: str | None = None) -> dict:
    """
    为每个指标编码构造语义密度高的查询文本，供 embedding 计算使用。

    优先级：
      1. 若 enhanced_query_path（默认 ENHANCED_QUERY_PATH）文件存在且该编码有增强文本 →
         使用 LLM 生成的 embedding 友好型描述（包含关键词、同义词、典型资料类型）
      2. 否则回退到原始公式："{code} {topic} {indicator}：{requirement[:500]}"

    返回：
        {code: query_text}  共 189 条，每条均为非空字符串。
    """
    import json as _json

    eq_path = enhanced_query_path if enhanced_query_path is not None else None
    if eq_path is None:
        # 向后兼容：使用 config.py 中的默认路径
        from config import ENHANCED_QUERY_PATH as _EQP
        eq_path = _EQP

    # 尝试加载 LLM 增强文本
    enhanced = {}
    if os.path.isfile(eq_path):
        try:
            with open(eq_path, "r", encoding="utf-8") as f:
                enhanced = _json.load(f)
            print(f"  ✓ 加载 {len(enhanced)} 条增强查询文本（{eq_path}）")
        except Exception as e:
            print(f"  [警告] 加载增强查询文本失败：{e}，将使用原始公式")

    queries = {}
    enhanced_count = 0
    fallback_count = 0

    for code, info in indicator_details.items():
        # 优先使用增强文本
        enh_text = enhanced.get(code, "").strip()
        if enh_text:
            # 增强文本前加上编码，确保编码始终出现在查询中
            queries[code] = f"{code} {enh_text}"
            enhanced_count += 1
        else:
            # 回退到原始公式
            topic       = info.get("topic", "").strip()
            indicator   = info.get("indicator", "").strip()
            requirement = info.get("requirement", "").strip()[:500]

            parts = [p for p in [code, topic, indicator] if p]
            base  = " ".join(parts)

            if requirement:
                query = f"{base}：{requirement}"
            else:
                query = base

            queries[code] = query
            fallback_count += 1

    if enhanced:
        print(f"    增强文本：{enhanced_count} 条，原始公式回退：{fallback_count} 条")

    return queries


# ==============================================================================
# 阶段三 — 函数二：批量计算 Embedding（通用，供指标和 chunk 共用）
# ==============================================================================

def compute_embeddings(
    texts: list,
    api_key: str = None,
    base_url: str = None,
    model: str = None,
    batch_size: int = None,
    label: str = "",
    max_concurrent: int = None,
) -> list:
    """
    批量并发调用 DashScope TextEmbedding API，返回与输入等长的向量列表。

    自动检测 Instruct 前缀：
      - 文本含 "Instruct: ...\nQuery: ..." → text_type="query" + instruct 参数
      - 无此前缀 → text_type="document"
    （一个批次内所有文本的 text_type 须一致，由首条文本决定。）

    参数：
        texts      - 待 embedding 的文本列表
        api_key    - （兼容保留，不再使用）由 config.DASHSCOPE_API_KEY 提供
        base_url   - （兼容保留，不再使用）已改为直接 SDK 调用
        model      - （兼容保留，不再使用）由 config.EMBEDDING_MODEL 提供
        batch_size - 每批最多条数，默认从 config.EMBEDDING_BATCH_SIZE 读取
        label      - 进度打印前缀（如 "指标" 或 "chunk"）
        max_concurrent - 最大并发请求数，默认从 config.EMBEDDING_CONCURRENCY 读取

    返回：
        List[List[float]]，长度与 texts 相同。
        单批失败时对应位置填充零向量（填 []），打印警告。

    重试策略：每批最多重试 3 次，等待 1s / 2s / 4s（指数退避）。
    空列表输入时直接返回空列表，不发起任何 API 调用。
    """
    if not texts:
        return []

    # 默认值从 config 读取
    if batch_size is None:
        from config import EMBEDDING_BATCH_SIZE
        batch_size = EMBEDDING_BATCH_SIZE
    if max_concurrent is None:
        from config import EMBEDDING_CONCURRENCY
        max_concurrent = EMBEDDING_CONCURRENCY

    import re
    import dashscope
    from http import HTTPStatus
    from config import DASHSCOPE_API_KEY, EMBEDDING_MODEL, EMBEDDING_DIM
    from concurrent.futures import ThreadPoolExecutor, as_completed

    dashscope.api_key = DASHSCOPE_API_KEY

    # ── Instruct 前缀检测（由首条文本决定整批的 text_type）──
    INSTRUCT_PATTERN = re.compile(r'^Instruct:\s*(.+?)\nQuery:\s*(.+)', re.DOTALL)

    def parse_instruct(text: str):
        m = INSTRUCT_PATTERN.match(text)
        if m:
            return m.group(1).strip(), m.group(2).strip()
        return None, text

    instruct_str, _ = parse_instruct(texts[0])
    if instruct_str:
        text_type = "query"
        clean_texts = [parse_instruct(t)[1] for t in texts]
    else:
        text_type = "document"
        clean_texts = texts

    results = [None] * len(texts)
    total   = len(texts)
    n_batch = (total + batch_size - 1) // batch_size

    def _process_batch(batch_idx):
        """处理单个批次（带重试），返回 (batch_idx, embeddings_list)。"""
        start = batch_idx * batch_size
        end   = min(start + batch_size, total)
        batch = clean_texts[start:end]

        for attempt in range(3):
            try:
                kwargs = {
                    "model": EMBEDDING_MODEL,
                    "input": batch,
                    "dimension": EMBEDDING_DIM,
                    "text_type": text_type,
                    "output_type": "dense",
                }
                if instruct_str and text_type == "query":
                    kwargs["instruct"] = instruct_str

                resp = dashscope.TextEmbedding.call(**kwargs)

                if resp.status_code != HTTPStatus.OK:
                    raise RuntimeError(f"DashScope API 错误: {getattr(resp, 'message', str(resp))}")

                return batch_idx, [item["embedding"] for item in resp.output["embeddings"]]
            except Exception as e:
                wait = 2 ** attempt  # 1s, 2s, 4s
                tqdm.write(f"  [警告] 批次 {batch_idx + 1} 第 {attempt + 1} 次失败：{e}，"
                           f"{'重试' if attempt < 2 else '放弃'}（等待 {wait}s）")
                time.sleep(wait)

        # 全部重试失败，返回空向量
        return batch_idx, [[] for _ in batch]

    desc = f"  embedding{' ' + label if label else ''}"

    if max_concurrent <= 1:
        # 串行模式（兼容旧行为或调试用）
        pbar = tqdm(range(n_batch), desc=desc, unit="批",
                    ncols=80, dynamic_ncols=False)
        for b in pbar:
            start = b * batch_size
            end   = min(start + batch_size, total)
            pbar.set_postfix_str(f"{end}/{total} 条", refresh=False)
            _, embs = _process_batch(b)
            for j, emb in enumerate(embs):
                results[start + j] = emb
        pbar.close()
    else:
        # 并发模式
        pbar = tqdm(total=n_batch, desc=desc, unit="批",
                    ncols=80, dynamic_ncols=False)
        with ThreadPoolExecutor(max_workers=max_concurrent) as executor:
            futures = {
                executor.submit(_process_batch, b): b
                for b in range(n_batch)
            }
            for future in as_completed(futures):
                batch_idx, embs = future.result()
                start = batch_idx * batch_size
                for j, emb in enumerate(embs):
                    results[start + j] = emb
                end = min(start + batch_size, total)
                pbar.set_postfix_str(f"{end}/{total} 条", refresh=False)
                pbar.update(1)
        pbar.close()

    return results


# ==============================================================================
# 阶段三 — 函数三：构建指标向量库（ChromaDB，持久化）
# ==============================================================================

def build_indicator_collection(
    indicator_queries: dict,
    indicator_details: dict,
    api_key: str,
    base_url: str,
    model: str,
    persist_dir: str,
    company_name: str,
):
    """
    计算 189 个指标查询文本的 embedding，存入 ChromaDB 持久化 collection。

    Collection 名称："{company_name}_indicators"（如 "艾森股份_indicators"）。

    复用逻辑：
        若 collection 已存在且 count() == len(indicator_queries)，
        直接返回，跳过 embedding 计算（节省 API 调用）。
        否则删除旧 collection，重建。

    ChromaDB 存储格式：
        ids       = [编码, ...]          如 ["GA1", "GA2", ...]
        embeddings= [向量, ...]
        metadatas = [{"code": ..., "topic": ..., "indicator": ...}, ...]

    返回：已填充的 chromadb.Collection 对象，供阶段四 query() 使用。
    """
    try:
        import chromadb
    except ImportError:
        print("  [错误] 未安装 chromadb 库，请执行：pip install chromadb")
        return None

    os.makedirs(persist_dir, exist_ok=True)
    client = chromadb.PersistentClient(path=persist_dir)

    # ChromaDB collection 名只允许 [a-zA-Z0-9._-]，且必须以字母数字开头结尾
    # 中文公司名转为 ASCII slug，统一加 "esg-" 前缀确保合法
    import re as _re
    _slug = _re.sub(r'[^a-zA-Z0-9._-]', '-', company_name).strip('-').strip('.')
    _slug = _re.sub(r'-{2,}', '-', _slug)  # 连续减号合并
    if not _slug:
        _slug = "default"
    collection_name = f"esg-{_slug}-indicators"
    expected_count  = len(indicator_queries)

    # 复用检测：collection 存在且条数一致 → 直接返回
    try:
        existing = client.get_collection(collection_name)
        if existing.count() == expected_count:
            print(f"  ✓ 指标 embedding 已缓存（{expected_count} 条），跳过重建")
            return existing
        else:
            print(f"  [提示] 缓存条数不一致（{existing.count()} vs {expected_count}），重建中...")
            client.delete_collection(collection_name)
    except Exception:
        pass  # collection 不存在，直接创建

    collection = client.create_collection(
        name     = collection_name,
        metadata = {"hnsw:space": "cosine"},  # 余弦相似度
    )

    codes  = list(indicator_queries.keys())
    texts  = [indicator_queries[c] for c in codes]

    # 指标 query 侧加 instruct 前缀（Qwen3-Embedding 推荐用法）
    # document/chunk 侧不加，保持非对称结构
    if EMBEDDING_INSTRUCT:
        texts = [f"Instruct: {EMBEDDING_INSTRUCT}\nQuery: {t}" for t in texts]

    metas  = [
        {
            "code":      c,
            "topic":     indicator_details.get(c, {}).get("topic", ""),
            "indicator": indicator_details.get(c, {}).get("indicator", ""),
        }
        for c in codes
    ]

    embeddings = compute_embeddings(texts, api_key, base_url, model, label="指标")

    # 过滤零向量（API 失败的条目），避免 ChromaDB 拒绝空向量
    valid_ids  = []
    valid_embs = []
    valid_meta = []
    for i, (c, emb) in enumerate(zip(codes, embeddings)):
        if emb:
            valid_ids.append(c)
            valid_embs.append(emb)
            valid_meta.append(metas[i])
        else:
            print(f"  [警告] 指标 {c} embedding 为空，已跳过（不存入 ChromaDB）")

    if valid_ids:
        collection.add(ids=valid_ids, embeddings=valid_embs, metadatas=valid_meta)

    print(f"  ✓ {len(valid_ids)} 个指标 embedding 已存入 ChromaDB")
    return collection


# ==============================================================================
# 阶段三 — 函数四：计算文本块 Embedding（内存驻留，不持久化）
# ==============================================================================

def embed_chunks(
    chunk_records: list,
    api_key: str,
    base_url: str,
    model: str,
) -> list:
    """
    为每个 chunk_record 追加 "embedding" 字段，返回增强后的新列表。
    原始 chunk_records 不被修改。

    向量化使用 get_text_for_embedding() 选择字段：
    - 表格 chunk 且有 table_summary：使用纯摘要（语义密度高）
    - 其他情况：使用 text 字段

    char_count == 0 的 chunk（空文本）embedding 设为 None，跳过 API 调用。

    返回：
        List[dict]，每条在原 ChunkRecord 基础上追加：
            "embedding": List[float] | None
    """
    # 分离有效 chunk（有文本）和空 chunk（无文本）
    valid_indices = [i for i, c in enumerate(chunk_records) if c.get("char_count", 0) > 0]
    # Phase 2：使用 get_text_for_embedding() 选择 embedding 输入文本
    valid_texts   = [get_text_for_embedding(chunk_records[i]) for i in valid_indices]

    # 批量计算有效 chunk 的 embedding
    if valid_texts:
        embeddings = compute_embeddings(valid_texts, api_key, base_url, model, label="chunk")
    else:
        embeddings = []

    # 组装结果，空 chunk 填 None
    emb_map = {valid_indices[j]: embeddings[j] for j in range(len(valid_indices))}

    result = []
    for i, chunk in enumerate(chunk_records):
        new_chunk = dict(chunk)
        new_chunk["embedding"] = emb_map.get(i, None)
        result.append(new_chunk)

    return result


# ==============================================================================
# 阶段四 — 函数一：批量语义检索
# ==============================================================================

def semantic_search_batch(
    chunk_records_with_emb: list,
    indicator_collection,             # chromadb.Collection | None
    top_k: int = EMBEDDING_TOP_K,
) -> list:
    """
    批量对所有 chunk 做 ChromaDB query，返回每条的语义 Top-K。

    参数：
        chunk_records_with_emb - 含 "embedding" 字段的 chunk 列表
        indicator_collection   - ChromaDB Collection 对象（cosine space）
        top_k                  - 每个 chunk 返回的候选指标数

    返回：
        List[List[Tuple[str, float]]]，与输入等长。
        每条为 [(code, similarity), ...]，按 score 降序。
        无效 embedding（None / 零向量 / 空列表）对应位置为 []。

    注意：
        - indicator_collection 为 None 时打印警告，全部返回 []
        - 所有 embedding 均无效时不调用 query()
        - 距离→相似度：similarity = 1.0 - distance（cosine space）
    """
    n = len(chunk_records_with_emb)
    results = [[] for _ in range(n)]

    if indicator_collection is None:
        print("  [警告] indicator_collection 为 None，跳过语义检索")
        return results

    # 筛选有效 embedding 的索引
    valid_indices = []
    valid_embeddings = []
    for i, rec in enumerate(chunk_records_with_emb):
        emb = rec.get("embedding")
        if emb is None:
            continue
        if not isinstance(emb, list) or len(emb) == 0:
            continue
        if sum(abs(v) for v in emb) <= 1e-9:
            continue
        valid_indices.append(i)
        valid_embeddings.append(emb)

    if not valid_indices:
        return results

    # 分批查询（ChromaDB/SQLite 对单次 query 的变量数有上限）
    QUERY_BATCH_SIZE = 5000
    all_ids_batched = {}       # j → ids list
    all_distances_batched = {} # j → distances list

    for batch_start in range(0, len(valid_embeddings), QUERY_BATCH_SIZE):
        batch_end = min(batch_start + QUERY_BATCH_SIZE, len(valid_embeddings))
        batch_embs = valid_embeddings[batch_start:batch_end]
        try:
            qr = indicator_collection.query(
                query_embeddings=batch_embs,
                n_results=top_k,
            )
        except Exception as e:
            print(f"  [警告] ChromaDB query 批次 {batch_start}-{batch_end} 失败：{e}")
            continue
        for local_j in range(len(batch_embs)):
            global_j = batch_start + local_j
            all_ids_batched[global_j] = qr["ids"][local_j] if qr.get("ids") else []
            all_distances_batched[global_j] = qr["distances"][local_j] if qr.get("distances") else []

    # 解析结果，按原始索引回填
    for j, orig_idx in enumerate(valid_indices):
        ids       = all_ids_batched.get(j, [])
        distances = all_distances_batched.get(j, [])
        topk = []
        for code, dist in zip(ids, distances):
            similarity = 1.0 - dist
            topk.append((code, similarity))
        results[orig_idx] = topk

    return results


# ==============================================================================
# 阶段四 — 函数二：一致性判定（纯函数）
# ==============================================================================

def classify_consistency(
    folder_code: str | None,
    semantic_topk: list,               # [(code, score), ...]
    topn: int = CONSISTENCY_TOPN,
    extra_threshold: float = EXTRA_RELEVANCE_THRESHOLD,
    min_relevance: float = MIN_RELEVANCE_SCORE,
) -> tuple[str, str, str | None]:
    """
    纯函数，判定单个 chunk 的一致性状态。

    参数：
        folder_code     - 路径编码（None 表示兜底文件夹）
        semantic_topk   - 语义检索结果 [(code, score), ...]
        topn            - 路径编码需在 Top-N 内才算一致
        extra_threshold - 额外关联阈值（严格大于）
        min_relevance   - 最低相关度阈值（top1 score < 此值 → 低相关）

    返回：
        (status_emoji, status_desc, suggested_code)

    六分支判定树：
        1. semantic_topk 为空：
           - folder_code != None → ("❓", "有路径编码但无语义验证", folder_code)
           - folder_code == None → ("❓", "无任何证据", None)
        2. top1 score < min_relevance（低相关，与任何指标都不密切）
           → ("➖", "低相关", folder_code or topk[0][0])
        3. folder_code == None + 语义有结果
           → ("🔍", "无路径标签但语义命中", topk[0][0])
        4. folder_code ∈ top_n_codes
           - 有其他 code with score > extra_threshold → ("➕", "一致且有额外关联", folder_code)
           - 无 → ("✅", "一致", folder_code)
        5. folder_code ∉ top_n_codes
           → ("⚠️", "疑似错位", topk[0][0])
    """
    # 分支 1：语义无结果
    if not semantic_topk:
        if folder_code is not None:
            return ("❓", "有路径编码但无语义验证", folder_code)
        else:
            return ("❓", "无任何证据", None)

    # 分支 2：top1 score 低于最低相关度 → 低相关
    top1_score = semantic_topk[0][1] if semantic_topk else 0
    if top1_score < min_relevance:
        return ("➖", "低相关", folder_code or semantic_topk[0][0])

    # 分支 3：无路径标签但语义有结果
    if folder_code is None:
        return ("🔍", "无路径标签但语义命中", semantic_topk[0][0])

    # 提取 Top-N 编码
    top_n_codes = [code for code, _score in semantic_topk[:topn]]

    # 分支 3：folder_code 在 Top-N 内
    if folder_code in top_n_codes:
        # 检查是否有额外关联（其他编码 score > extra_threshold）
        has_extra = any(
            score > extra_threshold
            for code, score in semantic_topk
            if code != folder_code
        )
        if has_extra:
            return ("➕", "一致且有额外关联", folder_code)
        else:
            return ("✅", "一致", folder_code)

    # 分支 4：folder_code 不在 Top-N 内
    return ("⚠️", "疑似错位", semantic_topk[0][0])


# ==============================================================================
# 阶段四 — 函数三：主整合函数
# ==============================================================================

def align_chunks(
    chunk_records_with_emb: list,
    indicator_collection,
    top_k: int = EMBEDDING_TOP_K,
    topn: int = CONSISTENCY_TOPN,
    extra_threshold: float = EXTRA_RELEVANCE_THRESHOLD,
    min_relevance: float = MIN_RELEVANCE_SCORE,
) -> list:
    """
    主整合函数：调用 semantic_search_batch + classify_consistency。

    参数：
        chunk_records_with_emb - 含 embedding 的 chunk 列表
        indicator_collection   - ChromaDB Collection 对象
        top_k                  - 语义检索候选数
        topn                   - 一致性判定 Top-N
        extra_threshold        - 额外关联阈值
        min_relevance          - 最低相关度阈值

    返回：
        List[dict]，每条追加 4 个字段：
            "semantic_topk":    [(code, score), ...]
            "consistency":      状态 emoji
            "consistency_desc": 中文描述
            "suggested_code":   建议编码

    不修改原始 chunk_records_with_emb（浅拷贝每条记录）。
    """
    # 批量语义检索
    all_topk = semantic_search_batch(
        chunk_records_with_emb, indicator_collection, top_k=top_k
    )

    # 逐条判定一致性
    alignment_records = []
    for i, rec in enumerate(chunk_records_with_emb):
        new_rec = dict(rec)  # 浅拷贝，不修改原始记录
        topk = all_topk[i]
        status, desc, suggested = classify_consistency(
            folder_code=rec.get("folder_code"),
            semantic_topk=topk,
            topn=topn,
            extra_threshold=extra_threshold,
            min_relevance=min_relevance,
        )
        new_rec["semantic_topk"]    = topk
        new_rec["consistency"]      = status
        new_rec["consistency_desc"] = desc
        new_rec["suggested_code"]   = suggested
        alignment_records.append(new_rec)

    # 打印统计摘要
    print_phase4_summary(alignment_records)

    return alignment_records


# ==============================================================================
# 阶段四 — 辅助函数：统计摘要打印
# ==============================================================================

def print_phase4_summary(alignment_records: list) -> None:
    """
    打印阶段四统计摘要表（各状态数量和占比），按固定顺序排列。
    """
    total = len(alignment_records)
    if total == 0:
        print("  统计摘要：无记录")
        return

    # 固定顺序的状态列表
    status_order = [
        ("✅", "一致"),
        ("➕", "一致且有额外关联"),
        ("⚠️", "疑似错位"),
        ("🔍", "无路径标签但语义命中"),
        ("➖", "低相关"),
        ("❓", "无命中"),
    ]

    # 统计各状态数量
    counts = {}
    for rec in alignment_records:
        emoji = rec.get("consistency", "❓")
        counts[emoji] = counts.get(emoji, 0) + 1

    # 打印表格
    print("  统计摘要:")
    print("  ┌──────────────────────────────┬──────┬───────┐")
    print("  │ 状态                         │ 数量 │ 占比  │")
    print("  ├──────────────────────────────┼──────┼───────┤")
    for emoji, label in status_order:
        count = counts.get(emoji, 0)
        pct   = count / total * 100
        status_str = f"{emoji} {label}"
        print(f"  │ {status_str:<28} │ {count:>4} │ {pct:>5.1f}% │")
    print("  └──────────────────────────────┴──────┴───────┘")


# ==============================================================================
# 阶段五 — 对齐表列定义与排序
# ==============================================================================

ALIGNMENT_COLUMNS = [
    "chunk_id", "file_name", "file_path", "folder_code",
    "folder_topic", "folder_indicator", "semantic_top5",
    "consistency", "text_preview", "page_or_sheet",
    "chunk_index", "human_code", "human_note",
]

CONSISTENCY_SORT_ORDER = {"⚠️": 0, "🔍": 1, "➕": 2, "✅": 3, "➖": 4, "❓": 5}


# ==============================================================================
# 阶段五 — 函数一：输出对齐表 Excel
# ==============================================================================

def write_alignment_excel(
    alignment_records: list,
    indicator_details: dict,
    output_dir: str,
    target_folder: str,
    company_name: str,
    file_records: list | None = None,
) -> str:
    """
    将 alignment_records 转换为 13 列 DataFrame，排序、写入 Excel、格式化。
    若 file_records 提供，额外写入「未覆盖文件」sheet。

    参数：
        alignment_records  - 阶段四产出的对齐记录列表
        indicator_details  - 指标详情映射 {code: {"topic", "indicator", "requirement"}}
        output_dir         - Excel 输出目录
        target_folder      - 计算相对路径的基准目录
        company_name       - 公司名称（未来可用于文件名）

    返回：
        生成的 Excel 文件绝对路径。
    """
    import pandas as pd
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    import re as _re

    # openpyxl 禁止写入的非法字符（XML 1.0 不允许的控制字符）
    _ILLEGAL_CHARS_RE = _re.compile(
        r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
    )

    def _sanitize(text: str) -> str:
        """移除 openpyxl / Excel 不接受的非法字符。"""
        if not isinstance(text, str):
            return text
        return _ILLEGAL_CHARS_RE.sub('', text)

    # ── 步骤 1：构建行数据 ──
    rows = []
    for rec in alignment_records:
        folder_code = rec.get("folder_code")
        # 查表补 folder_topic / folder_indicator
        if folder_code and folder_code in indicator_details:
            detail = indicator_details[folder_code]
            folder_topic = detail.get("topic", "") or ""
            folder_indicator = detail.get("indicator", "") or ""
        else:
            folder_topic = ""
            folder_indicator = ""

        # 格式化 semantic_topk
        topk = rec.get("semantic_topk", []) or []
        semantic_top5 = ", ".join(
            f"{code}:{score:.2f}" for code, score in topk
        ) if topk else ""

        # text_preview：子块文本前 200 字，替换换行符
        text = rec.get("text", "") or ""
        text_preview = text[:200].replace("\n", " ").replace("\r", " ")

        # human_code 预填 suggested_code
        human_code = rec.get("suggested_code") or ""

        # file_path 转为相对路径
        abs_path = rec.get("file_path", "")
        try:
            rel_path = os.path.relpath(abs_path, target_folder)
        except ValueError:
            rel_path = abs_path

        rows.append({
            "chunk_id":         _sanitize(rec.get("chunk_id", "") or ""),
            "file_name":        _sanitize(rec.get("file_name", "") or ""),
            "file_path":        _sanitize(rel_path),
            "folder_code":      _sanitize(folder_code or ""),
            "folder_topic":     _sanitize(folder_topic),
            "folder_indicator": _sanitize(folder_indicator),
            "semantic_top5":    _sanitize(semantic_top5),
            "consistency":      rec.get("consistency", "") or "",
            "text_preview":     _sanitize(text_preview),
            "page_or_sheet":    _sanitize(rec.get("page_or_sheet", "") or ""),
            "chunk_index":      rec.get("chunk_index", 0),
            "human_code":       _sanitize(human_code),
            "human_note":       "",
        })

    # ── 步骤 2：构建 DataFrame ──
    df = pd.DataFrame(rows, columns=ALIGNMENT_COLUMNS)

    # ── 步骤 3：排序 ──
    df["_sort_key"] = df["consistency"].map(
        lambda x: CONSISTENCY_SORT_ORDER.get(x, 999)
    )
    df = df.sort_values("_sort_key", kind="mergesort").drop(columns="_sort_key").reset_index(drop=True)

    # ── 步骤 4：构造输出路径 ──
    os.makedirs(output_dir, exist_ok=True)
    filename = f"对齐表_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(output_dir, filename)

    # ── 步骤 5：pandas 写 Excel ──
    df.to_excel(output_path, index=False, engine="openpyxl")

    # ── 步骤 6：openpyxl 后处理 ──
    wb = load_workbook(output_path)
    ws = wb.active

    # 6a. 冻结首行
    ws.freeze_panes = "A2"

    # 6b. 表头行加粗 + 浅灰背景
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 6c. 条件格式：⚠️ 行浅红，🔍 行浅黄，➖ 行浅灰
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    gray_font = Font(color="808080")
    # consistency 列是第 8 列（H 列，1-indexed col=8）
    consistency_col_idx = ALIGNMENT_COLUMNS.index("consistency") + 1  # 1-indexed

    for row_idx in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=row_idx, column=consistency_col_idx).value or "")
        if "⚠️" in cell_value or cell_value == "⚠️":
            fill, font = red_fill, None
        elif "🔍" in cell_value or cell_value == "🔍":
            fill, font = yellow_fill, None
        elif "➖" in cell_value or cell_value == "➖":
            fill, font = gray_fill, gray_font
        else:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
            if font:
                ws.cell(row=row_idx, column=col_idx).font = font

    # 6d. 自动列宽（CJK 字符按 1.7 倍估算，上限 60 字符）
    def _estimate_width(text: str) -> float:
        """估算显示宽度：CJK 字符按 1.7 倍，其他按 1 倍。"""
        import unicodedata
        width = 0.0
        for ch in str(text):
            if unicodedata.east_asian_width(ch) in ('W', 'F'):
                width += 1.7
            else:
                width += 1.0
        return width

    for col_cells in ws.columns:
        max_width = 0.0
        for cell in col_cells:
            try:
                cell_width = _estimate_width(str(cell.value or ""))
                if cell_width > max_width:
                    max_width = cell_width
            except Exception:
                pass
        # +2 余量，上限 60
        adjusted = min(max_width + 2, 60)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted

    # 6e. text_preview 列设置自动换行
    text_preview_col_idx = ALIGNMENT_COLUMNS.index("text_preview") + 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=text_preview_col_idx).alignment = Alignment(wrap_text=True)

    # ── 步骤 7：未覆盖文件 sheet ──
    if file_records:
        relevant_statuses = {"✅", "➕", "⚠️", "🔍"}
        covered_files = set(
            rec.get("file_path", "")
            for rec in alignment_records
            if rec.get("consistency", "") in relevant_statuses
        )
        all_file_map = {fr.get("file_path", ""): fr for fr in file_records}
        uncovered_paths = set(all_file_map.keys()) - covered_files

        if uncovered_paths:
            ws2 = wb.create_sheet(title="未覆盖文件")
            ws2_headers = ["文件名", "相对路径", "文件夹编码", "chunk数", "说明"]
            for col_idx, h in enumerate(ws2_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font

            # 统计每个文件的 chunk 数
            file_chunk_counts = {}
            for rec in alignment_records:
                fp = rec.get("file_path", "")
                file_chunk_counts[fp] = file_chunk_counts.get(fp, 0) + 1

            uncovered_rows = []
            for fp in sorted(uncovered_paths):
                fr = all_file_map[fp]
                try:
                    rel = os.path.relpath(fp, target_folder)
                except ValueError:
                    rel = fp
                uncovered_rows.append({
                    "文件名": fr.get("file_name", ""),
                    "相对路径": rel,
                    "文件夹编码": fr.get("folder_code", "") or "",
                    "chunk数": file_chunk_counts.get(fp, 0),
                    "说明": "所有chunk均低相关（top1 < {:.2f}）".format(MIN_RELEVANCE_SCORE)
                           if file_chunk_counts.get(fp, 0) > 0
                           else "提取失败或无有效内容",
                })

            for row_idx, row in enumerate(uncovered_rows, 2):
                for col_idx, h in enumerate(ws2_headers, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=row[h])

            # 自动列宽
            for col_idx, h in enumerate(ws2_headers, 1):
                max_len = len(h) * 2  # 中文标题
                for row_idx in range(2, len(uncovered_rows) + 2):
                    val = str(ws2.cell(row=row_idx, column=col_idx).value or "")
                    max_len = max(max_len, len(val))
                ws2.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 60)

            ws2.freeze_panes = "A2"

    # ── 步骤 8：保存并返回 ──
    wb.save(output_path)
    return output_path


# ==============================================================================
# 阶段五 — 函数二：打印摘要
# ==============================================================================

def print_phase5_summary(
    output_path: str,
    alignment_records: list,
    file_records: list,
) -> None:
    """打印阶段五统计摘要，含文件覆盖率分析。"""
    review_statuses = {"⚠️", "🔍", "➕"}
    count = sum(
        1 for rec in alignment_records
        if rec.get("consistency", "") in review_statuses
    )
    print(f"  ✓ 已生成: {output_path}")
    print(f"  ✓ 需人工审查: {count} 行（⚠️ + 🔍 + ➕）")

    # 文件覆盖率分析：有效 chunk（非 ➖ 非 ❓）覆盖了多少文件
    relevant_statuses = {"✅", "➕", "⚠️", "🔍"}
    covered_files = set(
        rec.get("file_path", "")
        for rec in alignment_records
        if rec.get("consistency", "") in relevant_statuses
    )
    all_files = set(fr.get("file_path", "") for fr in file_records)
    uncovered = all_files - covered_files
    covered_count = len(all_files) - len(uncovered)

    print(f"  ✓ 文件覆盖率: {covered_count}/{len(all_files)}"
          f"（{covered_count / len(all_files) * 100:.1f}%）")
    if uncovered:
        print(f"  ⚠ 未覆盖文件（所有 chunk 均低相关）: {len(uncovered)} 个")
        # 按相对路径排序，取前 10 个展示
        uncovered_rel = sorted(
            os.path.relpath(fp, os.path.dirname(output_path).replace("/processed", "/processed/模拟甲方整理后资料"))
            if fp else fp
            for fp in uncovered
        )
        for fp in uncovered_rel[:10]:
            print(f"    - {fp}")
        if len(uncovered_rel) > 10:
            print(f"    ... 及其余 {len(uncovered_rel) - 10} 个（详见对齐表「未覆盖文件」sheet）")

    print()
    print("═" * 39)


# ==============================================================================
# 主函数
# ==============================================================================

# 各文件类型对应的提取器（按扩展名路由）
_EXTRACTOR_MAP = {
    ".pdf":  extract_pdf,
    ".docx": extract_docx,
    ".doc":  extract_doc,
    ".xlsx": extract_xlsx,
    ".xls":  extract_xls,
    ".pptx": extract_pptx,
    ".ppt":  extract_ppt,
    ".jpg":  extract_image,
    ".jpeg": extract_image,
    ".png":  extract_image,
}


def main():
    args = parse_args()

    # ── 路径初始化（多企业支持，向后兼容）────────────────────────────────────
    paths = get_paths(args.project_dir)
    # 确保 processed 目录存在（新企业首次运行时自动创建）
    paths.processed_dir.mkdir(parents=True, exist_ok=True)

    # company_name 从 project_dir 目录名提取（get_paths 已在 project_dir=None 时抛出 ValueError）
    company_name = paths.project_dir.name

    # ── 缓存重建控制（--rebuild 替代旧的 FORCE_REEXTRACT / FORCE_RECHUNK）────
    rebuild = args.rebuild
    if rebuild == "all":
        rebuild = "extract"

    force_reextract = (rebuild == "extract")
    force_rechunk   = (rebuild in ("extract", "chunk"))
    force_reembed   = (rebuild in ("extract", "chunk", "embedding"))

    print_header(company_name)
    print_config_summary(paths, company_name, rebuild=args.rebuild)

    if rebuild:
        _cleanup_caches(paths, rebuild)

    # Web UI 进度追踪（CLI 无 --tracker 时为 NullTracker，零开销）
    tracker = get_tracker(args, "align_evidence")

    # 耗时统计
    from stage_timer import StageTimer
    timer = StageTimer()

    # 加载 VLM 缓存
    load_vlm_cache(str(paths.vlm_cache))

    # ── 阶段一：加载清单映射 + 扫描文件 ──────────────────────────────────
    timer.start("阶段一：加载清单映射")
    tracker.set_stage("Load manifest")
    print("[阶段一] 加载清单映射...")

    esg_mapping, _stats   = load_esg_mapping_from_reference_excel(str(paths.checklist_xlsx))
    indicator_details     = load_indicator_details(str(paths.checklist_xlsx))
    file_records          = scan_target_files(str(paths.materials_dir), esg_mapping)

    print_phase1_summary(esg_mapping, indicator_details, file_records)

    # 加载增强查询文本，注入给 VLM 作为图片分类上下文
    import json as _json
    _enhanced_for_vlm = {}
    if os.path.isfile(str(paths.enhanced_query)):
        try:
            with open(str(paths.enhanced_query), "r", encoding="utf-8") as _f:
                _enhanced_for_vlm = _json.load(_f)
            print(f"  ✓ 已加载增强查询文本用于 VLM 上下文：{len(_enhanced_for_vlm)} 条")
        except Exception:
            pass
    configure_vlm_context(_enhanced_for_vlm)

    # ── 阶段 2a：文本提取 → sections（extractors.py）───────────────────
    timer.start("阶段 2a：文本提取")
    print()
    print("[阶段 2a] 文本提取...")
    tracker.set_stage("Text extraction", total=len(file_records))

    all_sections = None
    if not force_reextract:
        all_sections = load_sections_cache(str(paths.section_cache))
        if all_sections is not None:
            total_secs = sum(len(v) for v in all_sections.values())
            print(f"  ✓ 从缓存加载 {len(all_sections)} 个文件的 {total_secs} 个 sections"
                  f"（跳过重新提取）")
            print(f"    缓存路径：{paths.section_cache}")

    if all_sections is None:
        if force_reextract:
            print(f"  [缓存重建] --rebuild {args.rebuild}，重新提取")
            # 清除 VLM 缓存（确保图片也重新处理）
            if os.path.isfile(str(paths.vlm_cache)):
                os.remove(str(paths.vlm_cache))
                print(f"  [强制重提取] 已删除 VLM 缓存：{paths.vlm_cache}")
        all_sections = {}
        skipped = 0

        # ── 并发提取（文件级） ─────────────────────────────────────────
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from config import EXTRACT_CONCURRENCY
        import threading

        _sections_lock = threading.Lock()
        completed = 0

        def _extract_one(fr):
            """提取单个文件，返回 (relative_path, sections) 或 (None, None)。"""
            sections = extract_sections(fr, img_base_dir=str(paths.sdk_image_dir))
            if sections is None:
                return (None, None)
            rel_path = fr.get("relative_path", fr["file_name"])
            return (rel_path, sections)

        print(f"  并发提取: {EXTRACT_CONCURRENCY} 线程")
        pbar = tqdm(total=len(file_records), desc="  提取文件", unit="个",
                    ncols=80, dynamic_ncols=False)

        with ThreadPoolExecutor(max_workers=EXTRACT_CONCURRENCY) as pool:
            futures = {
                pool.submit(_extract_one, fr): fr
                for fr in file_records
            }
            for future in as_completed(futures):
                fr = futures[future]
                fname = fr["file_name"]
                try:
                    rel_path, sections = future.result()
                    if rel_path is None:
                        skipped += 1
                    else:
                        with _sections_lock:
                            all_sections[rel_path] = sections
                except Exception as e:
                    print(f"  [警告] 提取异常：{fname} — {e}")
                with _sections_lock:
                    completed += 1
                    pbar.update(1)
                    pbar.set_postfix_str(fname[:30], refresh=False)
                tracker.advance(1, detail=fname[:30])

        pbar.close()

        total_secs = sum(len(v) for v in all_sections.values())
        print(f"  ✓ 共提取 {total_secs} 个 sections（{len(all_sections)} 个文件）"
              + (f"（跳过 {skipped} 个不支持格式）" if skipped else ""))
        save_sections_cache(all_sections, str(paths.section_cache))
        save_vlm_cache(str(paths.vlm_cache))  # 持久化 VLM 分类结果

    # ── 阶段 2b：分块 → chunks ──────────────────────────────────────────
    timer.start("阶段 2b：文本分块")
    print()
    print("[阶段 2b] 文本分块...")
    tracker.set_stage("Chunking", total=len(all_sections))

    # 扩展名 → chunk_params 文件类型
    _EXT_TYPE_MAP = {
        ".pdf": "pdf", ".docx": "docx", ".doc": "docx",
        ".xlsx": "xlsx", ".xls": "xlsx",
        ".pptx": "pptx", ".ppt": "pptx",
        ".jpg": "image", ".jpeg": "image", ".png": "image",
    }

    chunk_records = None
    all_parents = {}
    if not force_reextract and not force_rechunk:
        cached_data = load_chunks_cache(str(paths.chunk_cache))
        if cached_data is not None:
            chunk_records = cached_data.get("chunks", [])
            all_parents = cached_data.get("parents", {})
            print(f"  ✓ 从缓存加载 {len(chunk_records)} 个文本块（跳过重新分块）")
            print(f"    缓存路径：{paths.chunk_cache}")
            print(f"    如需强制重新分块，请使用 --rebuild chunk")

    if chunk_records is None:
        if force_rechunk:
            print(f"  [缓存重建] --rebuild {args.rebuild}，重新分块")
        chunk_records = []
        all_parents = {}
        fr_by_relpath = {fr.get("relative_path", fr["file_name"]): fr
                         for fr in file_records}
        for rel_path, sections in all_sections.items():
            fr = fr_by_relpath.get(rel_path)
            if fr is None:
                continue
            file_type = _EXT_TYPE_MAP.get(fr["extension"], "pdf")
            _max, _min = chunk_params(file_type)
            # Phase 1：make_chunks_from_sections 返回 {"parents": {...}, "chunks": [...]}
            result = make_chunks_from_sections(sections, fr,
                                               max_size=_max, min_size=_min)
            all_parents.update(result["parents"])
            chunk_records.extend(result["chunks"])

        # 统计表格 chunk 数量
        table_chunk_count = sum(1 for c in chunk_records if c.get("is_table"))
        print(f"  ✓ 共生成 {len(chunk_records)} 个文本块（含 {table_chunk_count} 个表格块）")
        save_chunks_cache({"parents": all_parents, "chunks": chunk_records}, str(paths.chunk_cache))

    # ── 阶段 2c：表格摘要生成（Phase 2） ─────────────────────────────────────
    timer.start("阶段 2c：表格摘要")
    print()
    print("[阶段 2c] 表格摘要生成...")
    tracker.set_stage("Table summaries")

    if ENABLE_TABLE_SUMMARY:
        # 检查是否有表格 chunk 需要生成摘要
        table_chunks_without_summary = [
            c for c in chunk_records
            if c.get("is_table") and not c.get("table_summary")
        ]

        if table_chunks_without_summary:
            print(f"  共 {len(table_chunks_without_summary)} 个表格 chunk 需要生成摘要...")

            # 异步调用 LLM 生成摘要
            import asyncio
            from table_summarizer import generate_table_summaries

            chunk_records = asyncio.run(
                generate_table_summaries(chunk_records, all_parents, cache_path=str(paths.table_summary_cache))
            )

            # 更新缓存（摘要生成后 text 字段已更新）
            save_chunks_cache({"parents": all_parents, "chunks": chunk_records}, str(paths.chunk_cache))

            summary_count = sum(1 for c in chunk_records
                               if c.get("is_table") and c.get("table_summary"))
            print(f"  ✓ 已生成 {summary_count} 个表格摘要")
        else:
            table_count = sum(1 for c in chunk_records if c.get("is_table"))
            if table_count > 0:
                print(f"  ✓ 所有 {table_count} 个表格 chunk 已有摘要（跳过生成）")
            else:
                print("  跳过（无表格 chunk）")
    else:
        print("  跳过（ENABLE_TABLE_SUMMARY=False）")

    # ── 阶段三：构建向量库 ────────────────────────────────────────────────
    timer.start("阶段三：构建向量库")
    print()
    print("[阶段三] 构建向量库...")
    tracker.set_stage("Build embeddings")

    # 3a：构建指标查询文本
    indicator_queries = build_indicator_queries(indicator_details, str(paths.enhanced_query))
    print(f"  ✓ 构建 {len(indicator_queries)} 条指标查询文本")

    # 3b + 3c：计算指标 embedding，存入 ChromaDB（含复用检测）
    indicator_collection = build_indicator_collection(
        indicator_queries,
        indicator_details,
        OPENAI_API_KEY,
        OPENAI_BASE_URL,
        EMBEDDING_MODEL,
        str(paths.chroma_dir),
        company_name,
    )

    # 3d：计算 chunk embedding（含 .npz 缓存复用）
    print(f"  文本块 embedding...（共 {len(chunk_records)} 个）")

    emb_list = None
    if not force_reembed:
        emb_list = load_emb_cache(str(paths.emb_cache), len(chunk_records))

    if emb_list is not None:
        chunk_records_with_emb = [
            {**rec, "embedding": emb}
            for rec, emb in zip(chunk_records, emb_list)
        ]
        valid_emb_count = sum(1 for e in emb_list if e is not None)
        print(f"  ✓ 从缓存加载 {len(emb_list)} 个 embedding"
              f"（{valid_emb_count} 有效，跳过重新计算）")
        print(f"    缓存路径：{paths.emb_cache}")
    else:
        if force_reembed:
            print(f"  [缓存重建] --rebuild {args.rebuild}，重算 embedding")
        chunk_records_with_emb = embed_chunks(
            chunk_records,
            OPENAI_API_KEY,
            OPENAI_BASE_URL,
            EMBEDDING_MODEL,
        )
        valid_emb_count = sum(
            1 for c in chunk_records_with_emb if c.get("embedding") is not None
        )
        print(f"  ✓ 全部 embedding 完成（{valid_emb_count} 个有效，"
              f"{len(chunk_records_with_emb) - valid_emb_count} 个空文本跳过）")
        save_emb_cache(chunk_records_with_emb, str(paths.emb_cache))

    # ── 阶段四：语义检索 + 一致性判断 ──────────────────────────────────
    timer.start("阶段四：语义检索")
    print()
    print("[阶段四] 语义检索与一致性判断...")
    tracker.set_stage("Semantic search")

    alignment_records = align_chunks(
        chunk_records_with_emb,
        indicator_collection,
        top_k=EMBEDDING_TOP_K,
        topn=CONSISTENCY_TOPN,
        extra_threshold=EXTRA_RELEVANCE_THRESHOLD,
        min_relevance=MIN_RELEVANCE_SCORE,
    )

    print(f"  ✓ 完成 {len(alignment_records)} 个文本块的对齐判断")

    # ── 阶段五：输出对齐表 ────────────────────────────────────────────────
    timer.start("阶段五：输出对齐表")
    print()
    print("[阶段五] 输出对齐表...")
    tracker.set_stage("Output Excel")

    output_path = write_alignment_excel(
        alignment_records,
        indicator_details,
        str(paths.processed_dir),
        str(paths.materials_dir),
        company_name,
        file_records=file_records,
    )

    print_phase5_summary(output_path, alignment_records, file_records)
    timer.report()
    tracker.complete()


if __name__ == "__main__":
    main()
